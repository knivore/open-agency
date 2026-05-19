from __future__ import annotations

import ast
import json
import os
from PIL import Image as PILImage, ImageOps
from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Any

IMAGE_WIDTH = 700
IMAGE_HEIGHT = 467
WIDTH_FUDGE_FACTOR = 0.142857
HEIGHT_FUDGE_FACTOR = 0.75


class ExcelImageInput(BaseModel):
    sheet_name: str = Field(..., description="The excel sheet name where the image will be embedded in")
    excel_file_path: str = Field(..., description="The Excel file path")
    image_path: str = Field(..., description="The image path of the image that will be embedded in the excel file")
    serial_number: int = Field(..., description="The serial number of testcase that will offset the rows in the excel")
    header_title: str | None = Field(default=None, description="The header text you want to add to the Excel file")


class ExcelJSONInput(BaseModel):
    sheet_name: str = Field(..., description="The excel sheet name where the json text will be embedded in")
    excel_file_path: str = Field(..., description="The Excel file path")
    json_file_path: str = Field(..., description="The JSON data file path")
    serial_number: int = Field(..., description="The serial number of testcase that will offset the rows in the excel")


class ExcelTextInput(BaseModel):
    sheet_name: str = Field(..., description="The excel sheet name where the text will be embedded in")
    excel_file_path: str = Field(..., description="The Excel file path")
    text_file_path: str = Field(..., description="The text data file path")
    serial_number: int = Field(..., description="The serial number of testcase that will offset the rows in the excel")
    header_title: str | None = Field(default=None, description="The header text you want to add to the Excel file")


def _column_name(index: int) -> str:
    name = ""
    while index >= 0:
        name = chr(index % 26 + ord("A")) + name
        index = index // 26 - 1
    return name


def _column_index(column: str) -> int:
    index = 0
    for char in column.upper():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _ensure_header_column(ws: Worksheet, title: str, offset: int, current_prefix: str | None = None) -> str:
    for col_idx, col in enumerate(ws.iter_cols(min_row=offset, max_row=offset, min_col=1, max_col=ws.max_column),
                                  start=1):
        cell = col[0]
        if isinstance(cell, MergedCell):
            cell = ws.cell(row=cell.row, column=cell.column)
        if cell.value == title:
            return _column_name(col_idx - 1)

    if current_prefix is not None:
        cell_prefix = current_prefix
    else:
        first_cell = ws.cell(row=offset, column=1).value
        next_index = 0 if ws.max_column <= 1 and first_cell in (None, "") else ws.max_column
        cell_prefix = _column_name(next_index)
    header_cell = f"{cell_prefix}{offset}"
    ws[header_cell] = title
    ws[header_cell].font = Font(bold=True)
    return cell_prefix


def _row_offset_for(row_offset: dict[str, Any] | None, sheet_name: str) -> int:
    return int((row_offset or {}).get(sheet_name, 0)) + 1


def _open_locked_workbook(excel_file_path: str):
    lock = FileLock(f"{excel_file_path}.lock", timeout=60)
    return lock


def write_excel_text(
        text_file_path: str,
        sheet_name: str,
        excel_file_path: str,
        serial_number: int,
        *,
        header_title: str | None = None,
        row_offset: dict[str, Any] | None = None,
        **_: Any,
) -> Any:
    title = header_title or "Text Output"
    offset = _row_offset_for(row_offset, sheet_name)
    try:
        with _open_locked_workbook(excel_file_path):
            workbook = load_workbook(excel_file_path, data_only=True)
            worksheet = workbook[sheet_name]
            with open(text_file_path, "r", encoding="utf-8") as handle:
                text_data = handle.read()

            cell_prefix = _ensure_header_column(worksheet, title, offset)
            row_col = f"{cell_prefix}{serial_number + offset}"
            worksheet[row_col].value = text_data
            worksheet[row_col].alignment = Alignment(wrap_text=True)
            worksheet.column_dimensions[cell_prefix].width = 50
            workbook.save(excel_file_path)
            return "{'Success Message': 'Text uploaded to Excel and successfully saved.'}"
    except FileNotFoundError:
        return {"Error Message": f"Excel file not found: {excel_file_path}"}
    except KeyError:
        return {"Error Message": f"Sheet name '{sheet_name}' not found in the workbook"}
    except Timeout:
        return {"Error Message": "File lock timeout exceeded. Could not access the Excel file."}
    except Exception as exc:
        return {"Error Message": f"Failed to save workbook: {exc}"}


def _read_json_file(file_path: str) -> dict[str, Any] | None:
    if not os.path.exists(file_path):
        return None
    try:
        with open(file_path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except json.JSONDecodeError:
        try:
            with open(file_path, "r", encoding="utf-8") as handle:
                parsed = ast.literal_eval(handle.read())
                return parsed if isinstance(parsed, dict) else {"content": parsed}
        except (ValueError, SyntaxError, FileNotFoundError):
            return None


def write_excel_json(
        json_file_path: str,
        sheet_name: str,
        excel_file_path: str,
        serial_number: int,
        *,
        row_offset: dict[str, Any] | None = None,
        **_: Any,
) -> Any:
    offset = _row_offset_for(row_offset, sheet_name)
    text_data = _read_json_file(json_file_path)
    if text_data is None:
        return {"Error Message": f"Invalid or missing JSON file: {json_file_path}"}

    try:
        with _open_locked_workbook(excel_file_path):
            workbook = load_workbook(excel_file_path, data_only=True)
            worksheet = workbook[sheet_name]
            current_prefix: str | None = None
            for key, value in text_data.items():
                current_prefix = _ensure_header_column(worksheet, key, offset, current_prefix)
                row_col = f"{current_prefix}{serial_number + offset}"
                worksheet[row_col].value = value
                worksheet[row_col].alignment = Alignment(wrap_text=True)
                worksheet.column_dimensions[current_prefix].width = 50
                current_prefix = _column_name(_column_index(current_prefix) + 1)

            workbook.save(excel_file_path)
            return "{'Success Message': 'JSON uploaded to Excel and successfully saved.'}"
    except FileNotFoundError:
        return {"Error Message": f"Excel file not found: {excel_file_path}"}
    except KeyError:
        return {"Error Message": f"Sheet name '{sheet_name}' not found in the workbook"}
    except Timeout:
        return {"Error Message": "File lock timeout exceeded. Could not access the Excel file."}
    except Exception as exc:
        return {"Error Message": f"Failed to save workbook: {exc}"}


def _resize_image_to_temp(image_path: str) -> Path:
    image_file = Path(image_path)
    with PILImage.open(image_file) as img:
        img = ImageOps.contain(img, (IMAGE_WIDTH, IMAGE_HEIGHT), PILImage.Resampling.LANCZOS)
        temp_image_path = image_file.with_stem(f"{image_file.stem}_resized")
        img.save(temp_image_path, quality=95, optimize=True)
        return temp_image_path


def _insert_image(worksheet: Worksheet, cell_prefix: str, cell_row: int, image_path: str) -> str | None:
    image_file = Path(image_path)
    if not image_file.exists():
        return f"The image file at {image_path} does not exist."

    temp_image_path = _resize_image_to_temp(str(image_file))
    worksheet.column_dimensions[cell_prefix].width = IMAGE_WIDTH * WIDTH_FUDGE_FACTOR
    if not worksheet.row_dimensions[cell_row].height or worksheet.row_dimensions[
        cell_row].height < IMAGE_HEIGHT * HEIGHT_FUDGE_FACTOR:
        worksheet.row_dimensions[cell_row].height = IMAGE_HEIGHT * HEIGHT_FUDGE_FACTOR
    image_to_store = OpenPyxlImage(str(temp_image_path))
    image_to_store.anchor = f"{cell_prefix}{cell_row}"
    worksheet.add_image(image_to_store)
    return None


def write_excel_image(
        sheet_name: str,
        excel_file_path: str,
        image_path: str,
        serial_number: int,
        *,
        header_title: str | None = None,
        save_all_ss_ind: bool = False,
        row_offset: dict[str, Any] | None = None,
        **_: Any,
) -> Any:
    title = header_title or "Image Results"
    offset = _row_offset_for(row_offset, sheet_name)

    try:
        with _open_locked_workbook(excel_file_path):
            workbook = load_workbook(excel_file_path, data_only=True)
            worksheet = workbook[sheet_name]
            cell_prefix = _ensure_header_column(worksheet, title, offset)
            cell_row = serial_number + offset

            if save_all_ss_ind and os.path.isdir(image_path):
                for img_file in sorted(os.listdir(image_path)):
                    if not img_file.endswith((".png", ".jpg", ".jpeg")):
                        continue
                    full_image_path = os.path.join(image_path, img_file)
                    error = _insert_image(worksheet, cell_prefix, cell_row, full_image_path)
                    if error:
                        return {"Error Message": error}
                    cell_prefix = _column_name(_column_index(cell_prefix) + 1)
            else:
                full_image_path = image_path
                if not save_all_ss_ind and os.path.isdir(image_path):
                    image_files = sorted(
                        [entry for entry in os.listdir(image_path) if entry.endswith((".png", ".jpg", ".jpeg"))]
                    )
                    if not image_files:
                        return {"Error Message": f"No image files found in directory: {image_path}"}
                    full_image_path = os.path.join(image_path, image_files[-1])
                error = _insert_image(worksheet, cell_prefix, cell_row, full_image_path)
                if error:
                    return {"Error Message": error}

            workbook.save(excel_file_path)
            return "{'Success Message': 'Images uploaded to Excel and successfully saved.'}"
    except FileNotFoundError:
        return {"Error Message": f"Excel file not found: {excel_file_path}"}
    except KeyError:
        return {"Error Message": f"Sheet name '{sheet_name}' not found in the workbook"}
    except Timeout:
        return {"Error Message": "File lock timeout exceeded. Could not access the Excel file."}
    except Exception as exc:
        return {"Error Message": f"Failed to save workbook: {exc}"}


__all__ = [
    "ExcelImageInput",
    "ExcelJSONInput",
    "ExcelTextInput",
    "write_excel_image",
    "write_excel_json",
    "write_excel_text",
]
