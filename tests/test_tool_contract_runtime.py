from __future__ import annotations

import os
import asyncio
import json
import subprocess
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image

from app.api.context import create_test_api_context
from app.api.main import create_app
from app.core.config import reset_settings_cache
from app.core.onecli_http import ONECLI_BLOCKED_HEADER_NAMES, ONECLI_BLOCKED_QUERY_PARAM_NAMES
from app.domain import (
    AgentDefinition,
    ApprovalRequest,
    ApprovalStatus,
    ApprovalTargetType,
    ApprovalType,
    Conversation,
    ConversationChannelType,
    DocumentUploadMode,
    Execution,
    ExecutionArtifact,
    ExecutionEvent,
    ExecutionEventType,
    ExecutionStatus,
    MemoryRecord,
    MemoryScope,
    MemoryType,
    ModelProfileDefinition,
    OneCLIIdentityMapping,
    ScheduleDefinition,
    TaskDefinition,
    UploadedDocument,
    UserDefinition,
    WorkflowDefinition,
)
from app.services.main_agent_setup.service import MainAgentSetupConfig, MainAgentSetupService
from app.services.executions import ExecutionService
from app.tools.cli_discovery import list_builtin_tool_definitions
from app.tools.contracts.loader import load_contracts
from app.tools.contracts.registry import ToolContractRegistry, get_default_contract_registry
from app.tools.contracts.validator import ToolContractValidationError, validate_tool_input, validate_tool_output
from app.tools.definitions import get_tool_catalog_specs
from app.tools.policies.engine import PolicyEngine
from app.tools.implementations.http_integrations import execute_custom_api
from app.tools.registry import ToolRegistry
from app.tools.runtime.executor import ToolRuntimeExecutor
from app.tools.runtime.pr_payloads import build_dry_run_pr_payload
from app.tools.runtime.store import JsonlToolRunStore
from app.services.workflows import WorkflowService
from app.runtime.streaming.event_bus import RuntimeEventBus, set_default_runtime_event_bus


README_PATCH = """diff --git a/README.md b/README.md
--- a/README.md
+++ b/README.md
@@ -1,1 +1,1 @@
-hello
+hello world
"""


class _RequestsResponse:
    status_code = 200
    text = '{"ok": true}'
    headers = {"Content-Type": "application/json"}

    def json(self):
        return {"ok": True}


def _create_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    workbook.save(path)
    return path


def _workflow_definition(workflow_id: str = "workflow-contract") -> WorkflowDefinition:
    return WorkflowDefinition(
        id=workflow_id,
        name="Contract Workflow",
        description="Workflow visible to contract runtime tests.",
        entrypoint="node-1",
        metadata={"inputs": ["topic"]},
    )


async def _prepare_conversation_context(conversation_id: str = "conversation-contract"):
    context = create_test_api_context()
    await context.model_profile_repo.save(
        ModelProfileDefinition(id="profile-contract", name="Contract Model", provider="fake", model="fake-model")
    )
    await MainAgentSetupService(context).create_main_agent(
        MainAgentSetupConfig(
            agent_name="Main Agent",
            agent_description="Configured for contract runtime tests.",
            agent_instructions="Answer briefly.",
            model_profile_id="profile-contract",
            profile_id="main-agent-profile",
            agent_id="main-agent",
            workflow_id="main-workflow",
        )
    )
    await context.conversation_repo.create(
        Conversation(
            id=conversation_id,
            created_by_user_id="user-contract",
            channel_type=ConversationChannelType.API,
        )
    )
    return context


def _workflow_payload(workflow_id: str = "workflow-contract-proposal") -> dict:
    return {
        "id": workflow_id,
        "name": "Contract Proposal Workflow",
        "description": "Workflow proposed through the contract runtime.",
        "entrypoint": "node-1",
        "nodes": [
            {
                "id": "node-1",
                "name": "Entry",
                "node_type": "task",
                "task_id": "task-1",
                "config": {},
                "metadata": {},
            }
        ],
        "task_definitions": [
            {
                "id": "task-1",
                "name": "Task One",
                "description": "Do the work",
                "tool_ids": [],
                "depends_on_task_ids": [],
                "input_schema": {},
                "output_schema": {},
                "human_approval_required": False,
                "framework_hints": {"preferred_adapter": None, "adapter_config": {}, "metadata": {}},
                "metadata": {},
            }
        ],
        "versioning": {
            "version": "1.0.0",
            "revision": 1,
            "parent_version": None,
            "is_published": False,
            "labels": [],
        },
        "metadata": {
            "visible_to_main_agent": True,
            "mutable_by_main_agent": True,
        },
    }


class ToolContractRuntimeTests(unittest.TestCase):
    def test_sandbox_edit_contract_loads_and_validates_payloads(self):
        contracts = load_contracts()
        registry = ToolContractRegistry(contracts)
        contract = registry.get_contract("sandbox-edit")
        http_contract = registry.get_contract("agency.http.request")
        workflow_list_contract = registry.get_contract("agency.workflow.list")
        workflow_get_contract = registry.get_contract("agency.workflow.get")
        execution_list_contract = registry.get_contract("agency.execution.list")
        execution_get_contract = registry.get_contract("agency.execution.get")
        execution_events_contract = registry.get_contract("agency.execution.events")
        execution_artifacts_contract = registry.get_contract("agency.execution.artifacts")
        tool_get_contract = registry.get_contract("agency.tool.get")
        workflow_propose_create_contract = registry.get_contract("agency.workflow.propose-create")
        workflow_propose_update_contract = registry.get_contract("agency.workflow.propose-update")
        tool_propose_create_contract = registry.get_contract("agency.tool.propose-create")
        tool_propose_update_contract = registry.get_contract("agency.tool.propose-update")
        memory_list_contract = registry.get_contract("agency.memory.list")
        memory_remember_contract = registry.get_contract("agency.memory.remember")
        memory_update_contract = registry.get_contract("agency.memory.update")
        memory_delete_contract = registry.get_contract("agency.memory.delete")
        tool_list_contract = registry.get_contract("agency.tool.list")
        command_run_contract = registry.get_contract("agency.command.run")
        file_write_contract = registry.get_contract("agency.file.write-text")
        document_contract = registry.get_contract("agency.document.markdown-to-word")
        excel_text_contract = registry.get_contract("agency.excel.write-text")
        excel_json_contract = registry.get_contract("agency.excel.write-json")
        excel_image_contract = registry.get_contract("agency.excel.write-image")

        self.assertIsNotNone(contract)
        self.assertEqual(contract.name, "sandbox-edit")
        self.assertIsNotNone(http_contract)
        self.assertEqual(http_contract.name, "agency.http.request")
        self.assertIsNotNone(workflow_list_contract)
        self.assertEqual(workflow_list_contract.name, "agency.workflow.list")
        self.assertIsNotNone(workflow_get_contract)
        self.assertEqual(workflow_get_contract.name, "agency.workflow.get")
        self.assertIsNotNone(execution_list_contract)
        self.assertEqual(execution_list_contract.name, "agency.execution.list")
        self.assertIsNotNone(execution_get_contract)
        self.assertEqual(execution_get_contract.name, "agency.execution.get")
        self.assertIsNotNone(execution_events_contract)
        self.assertEqual(execution_events_contract.name, "agency.execution.events")
        self.assertIsNotNone(execution_artifacts_contract)
        self.assertEqual(execution_artifacts_contract.name, "agency.execution.artifacts")
        self.assertIsNotNone(tool_get_contract)
        self.assertEqual(tool_get_contract.name, "agency.tool.get")
        self.assertIsNotNone(workflow_propose_create_contract)
        self.assertEqual(workflow_propose_create_contract.name, "agency.workflow.propose-create")
        self.assertIsNotNone(workflow_propose_update_contract)
        self.assertEqual(workflow_propose_update_contract.name, "agency.workflow.propose-update")
        self.assertIsNotNone(tool_propose_create_contract)
        self.assertEqual(tool_propose_create_contract.name, "agency.tool.propose-create")
        self.assertIsNotNone(tool_propose_update_contract)
        self.assertEqual(tool_propose_update_contract.name, "agency.tool.propose-update")
        self.assertIsNotNone(memory_list_contract)
        self.assertEqual(memory_list_contract.name, "agency.memory.list")
        self.assertIsNotNone(memory_remember_contract)
        self.assertEqual(memory_remember_contract.name, "agency.memory.remember")
        self.assertIsNotNone(memory_update_contract)
        self.assertEqual(memory_update_contract.name, "agency.memory.update")
        self.assertIsNotNone(memory_delete_contract)
        self.assertEqual(memory_delete_contract.name, "agency.memory.delete")
        self.assertIsNotNone(tool_list_contract)
        self.assertEqual(tool_list_contract.name, "agency.tool.list")
        self.assertIsNotNone(command_run_contract)
        self.assertEqual(command_run_contract.name, "agency.command.run")
        self.assertIn("shell", command_run_contract.risk_labels)
        self.assertIn("local_privileged_execution", command_run_contract.risk_labels)
        self.assertIsNotNone(file_write_contract)
        self.assertEqual(file_write_contract.name, "agency.file.write-text")
        self.assertIsNotNone(document_contract)
        self.assertEqual(document_contract.name, "agency.document.markdown-to-word")
        self.assertIsNotNone(excel_text_contract)
        self.assertEqual(excel_text_contract.name, "agency.excel.write-text")
        self.assertIsNotNone(excel_json_contract)
        self.assertEqual(excel_json_contract.name, "agency.excel.write-json")
        self.assertIsNotNone(excel_image_contract)
        self.assertEqual(excel_image_contract.name, "agency.excel.write-image")
        static_contract_names = {item.name for item in contracts}
        self.assertIn("agency.workflow.run", static_contract_names)
        self.assertIn("agency.human.ask", static_contract_names)
        self.assertIn("agency.browser.open", static_contract_names)
        self.assertIn("agency.browser.click", static_contract_names)
        validate_tool_input(
            contract,
            {
                "repo": "/tmp/example",
                "ref": "main",
                "changes": [{"path": "README.md", "patch": README_PATCH}],
                "dryRun": True,
            },
        )
        with self.assertRaises(ToolContractValidationError):
            validate_tool_input(contract, {"repo": "/tmp/example"})
        with self.assertRaises(ToolContractValidationError):
            validate_tool_output(contract, {"verdict": "maybe", "dryRun": True, "timestamp": "now"})

    def test_default_contract_registry_covers_every_builtin_tool(self):
        contracts = {contract.name for contract in get_default_contract_registry().list_contracts()}
        builtins = list_builtin_tool_definitions()
        missing = [tool.id for tool in builtins if tool.id not in contracts]

        self.assertEqual(missing, [])
        self.assertIn("agency.speech.listen", contracts)
        self.assertIn("agency.speech.speak", contracts)
        self.assertIn("agency.speech.continue", contracts)
        self.assertIn("sandbox-edit", contracts)
        self.assertIn("agency.memory.catalog", contracts)
        self.assertIn("agency.graph.context", contracts)
        self.assertIn("agency.schedule.list", contracts)
        self.assertIn("agency.workflow.runtime-governance.get", contracts)
        self.assertIn("agency.main-agent.monitor.get", contracts)
        self.assertIn("agency.documents.list", contracts)
        self.assertIn("agency.tool.workspace.list", contracts)
        self.assertIn("agency.tool.workspace.publish", contracts)
        self.assertIn("agency.workflow.improvement-proposals", contracts)
        self.assertIn("agency.workflow.steering-approvals", contracts)
        self.assertIn("agency.workflow.document-links", contracts)
        self.assertIn("agency.workflow.document-summary.get", contracts)
        self.assertIn("agency.workflow.shared-memory.namespaces", contracts)
        self.assertIn("agency.workflow.shared-memory.namespace.memory.add", contracts)
        self.assertIn("agency.workflow.improvement-proposal.request-approval", contracts)
        self.assertIn("agency.workflow.steering-approval.request-approval", contracts)
        self.assertIn("agency.workflow.governance.audit", contracts)
        self.assertIn("agency.workflow.governance.repair", contracts)
        self.assertIn("agency.workflow.governance.remediate", contracts)
        self.assertIn("agency.workflow.governance.review-queue", contracts)
        self.assertIn("agency.workflow.governance.act", contracts)
        self.assertIn("agency.workflow.governance.document-suggest", contracts)
        self.assertIn("agency.workflow.governance.bundle", contracts)
        self.assertIn("agency.graph.search", contracts)
        self.assertIn("agency.graph.expand", contracts)
        self.assertIn("agency.graph.neighbors", contracts)
        self.assertIn("agency.graph.path", contracts)
        self.assertIn("agency.graph.summarize-subgraph", contracts)
        self.assertIn("agency.graph.working-set.create", contracts)
        self.assertIn("agency.graph.working-set.add", contracts)
        self.assertIn("agency.graph.working-set.remove", contracts)
        self.assertIn("agency.graph.working-set.summarize", contracts)
        self.assertIn("agency.graph.working-set.clear", contracts)
        self.assertIn("agency.graph.working-set.persist-context-pack", contracts)
        self.assertIn("agency.workflow.memory-links.add", contracts)
        graph_context_contract = get_default_contract_registry().get_contract("agency.graph.context")
        graph_persist_contract = get_default_contract_registry().get_contract(
            "agency.graph.working-set.persist-context-pack"
        )
        self.assertIn("graph", graph_context_contract.risk_labels)
        self.assertIn("read_only", graph_context_contract.risk_labels)
        self.assertIn("graph", graph_persist_contract.risk_labels)
        self.assertIn("memory", graph_persist_contract.risk_labels)
        self.assertIn("context_pack", graph_persist_contract.risk_labels)
        self.assertIn("mutation", graph_persist_contract.risk_labels)

    def test_tool_run_executes_contract_backed_workflow_navigation_tools(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-nav",
                    name="Workflow Navigation",
                    entrypoint="node-1",
                    metadata={
                        "visible_to_main_agent": True,
                        "runtime_governance": {"token_budget": {"run_total_tokens": 1234}},
                    },
                )
            )
            await context.schedule_repo.create(
                ScheduleDefinition(
                    id="schedule-nav",
                    name="Nightly Workflow",
                    workflow_id="workflow-nav",
                )
            )
            await context.uploaded_document_repo.create(
                UploadedDocument(
                    id="document-nav",
                    filename="workflow-notes.txt",
                    extracted_text="Approval checklist for workflow navigation.\nValidate output before completion.",
                    content_sha256="abc123",
                    text_characters=12,
                    estimated_tokens=3,
                    upload_mode=DocumentUploadMode.VECTOR,
                    scope="workflow",
                    workflow_id="workflow-nav",
                    created_by_user_id="user-contract",
                    metadata={
                        "upload_intelligence": {
                            "summary": "Checklist for workflow approval and validation.",
                            "document_kind": "policy_sop",
                            "recommended": {"scope": "workflow", "tags": ["approval", "checklist"]},
                        }
                    },
                )
            )
            await context.memory_repo.create(
                MemoryRecord(
                    id="memory-nav",
                    scope=MemoryScope.WORKFLOW,
                    workflow_id="workflow-nav",
                    created_by_user_id="user-contract",
                    content="Validation checklist should be shared across runs.",
                    summary="Reusable validation checklist.",
                    memory_type=MemoryType.DECISION,
                )
            )
            executor = ToolRuntimeExecutor(context=context)

            schedule_response = await executor.run_async(
                "agency.schedule.list",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            governance_response = await executor.run_async(
                "agency.workflow.runtime-governance.get",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            monitor_response = await executor.run_async(
                "agency.main-agent.monitor.get",
                {},
                actor="user-contract",
            )
            create_proposal_response = await executor.run_async(
                "agency.workflow.improvement-proposal.create",
                {
                    "workflow_id": "workflow-nav",
                    "title": "Tighten output validation",
                    "summary": "Add a validation step before completion.",
                    "proposed_change": {"type": "task_instruction_update", "scope": "output"},
                    "validation_plan": "Run the workflow against one known sample.",
                },
                actor="user-contract",
            )
            list_proposal_response = await executor.run_async(
                "agency.workflow.improvement-proposals",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            proposal_id = str(create_proposal_response.result["proposal"]["id"])
            update_proposal_response = await executor.run_async(
                "agency.workflow.improvement-proposal.update",
                {
                    "workflow_id": "workflow-nav",
                    "proposal_id": proposal_id,
                    "patch": {"status": "approved"},
                },
                actor="user-contract",
            )
            proposal_approval_response = await executor.run_async(
                "agency.workflow.improvement-proposal.request-approval",
                {"workflow_id": "workflow-nav", "proposal_id": proposal_id},
                actor="user-contract",
            )
            create_steering_response = await executor.run_async(
                "agency.workflow.steering-approval.create",
                {
                    "workflow_id": "workflow-nav",
                    "recommended_action": "request_replan",
                    "reason": "The workflow needs a safer execution route.",
                    "operator_parameters": {"target_task_id": "task-1"},
                },
                actor="user-contract",
            )
            list_steering_response = await executor.run_async(
                "agency.workflow.steering-approvals",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            approval_id = str(create_steering_response.result["approval"]["id"])
            update_steering_response = await executor.run_async(
                "agency.workflow.steering-approval.update",
                {
                    "workflow_id": "workflow-nav",
                    "approval_id": approval_id,
                    "patch": {"status": "approved"},
                },
                actor="user-contract",
            )
            steering_approval_response = await executor.run_async(
                "agency.workflow.steering-approval.request-approval",
                {"workflow_id": "workflow-nav", "approval_id": approval_id},
                actor="user-contract",
            )
            governance_audit_response = await executor.run_async(
                "agency.workflow.governance.audit",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            governance_repair_response = await executor.run_async(
                "agency.workflow.governance.repair",
                {
                    "workflow_id": "workflow-nav",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                    "action": "sync_status_from_approval",
                },
                actor="user-contract",
            )
            governance_review_queue_response = await executor.run_async(
                "agency.workflow.governance.review-queue",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            document_summary_response = await executor.run_async(
                "agency.workflow.document-summary.get",
                {"workflow_id": "workflow-nav", "document_id": "document-nav"},
                actor="user-contract",
            )
            add_document_link_response = await executor.run_async(
                "agency.workflow.document-link.add",
                {
                    "workflow_id": "workflow-nav",
                    "document_id": "document-nav",
                    "target_type": "improvement_proposal",
                    "target_id": proposal_id,
                    "summary": "Evidence for the validation recommendation.",
                },
                actor="user-contract",
            )
            list_document_links_response = await executor.run_async(
                "agency.workflow.document-links",
                {"workflow_id": "workflow-nav", "document_id": "document-nav"},
                actor="user-contract",
            )
            document_link_id = str(add_document_link_response.result["link"]["id"])
            delete_document_link_response = await executor.run_async(
                "agency.workflow.document-link.delete",
                {"workflow_id": "workflow-nav", "link_id": document_link_id},
                actor="user-contract",
            )
            create_namespace_response = await executor.run_async(
                "agency.workflow.shared-memory.namespace.create",
                {
                    "workflow_id": "workflow-nav",
                    "name": "Validation Pack",
                    "description": "Shared validation context for this workflow.",
                },
                actor="user-contract",
            )
            namespace_id = str(create_namespace_response.result["namespace"]["id"])
            list_namespaces_response = await executor.run_async(
                "agency.workflow.shared-memory.namespaces",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )
            add_namespace_memory_response = await executor.run_async(
                "agency.workflow.shared-memory.namespace.memory.add",
                {
                    "workflow_id": "workflow-nav",
                    "namespace_id": namespace_id,
                    "memory_id": "memory-nav",
                },
                actor="user-contract",
            )
            namespace_memories_response = await executor.run_async(
                "agency.workflow.shared-memory.namespace.memories",
                {
                    "workflow_id": "workflow-nav",
                    "namespace_id": namespace_id,
                },
                actor="user-contract",
            )
            remove_namespace_memory_response = await executor.run_async(
                "agency.workflow.shared-memory.namespace.memory.remove",
                {
                    "workflow_id": "workflow-nav",
                    "namespace_id": namespace_id,
                    "memory_id": "memory-nav",
                },
                actor="user-contract",
            )
            documents_response = await executor.run_async(
                "agency.documents.list",
                {"workflow_id": "workflow-nav"},
                actor="user-contract",
            )

            self.assertEqual(schedule_response.verdict, "ok")
            self.assertEqual(schedule_response.result["count"], 1)
            self.assertEqual(schedule_response.result["items"][0]["id"], "schedule-nav")

            self.assertEqual(governance_response.verdict, "ok")
            self.assertEqual(governance_response.result["runtime_governance"]["workflow_id"], "workflow-nav")
            self.assertEqual(
                governance_response.result["runtime_governance"]["token_budget"]["run_total_tokens"],
                1234,
            )

            self.assertEqual(monitor_response.verdict, "ok")
            self.assertIn("summary", monitor_response.result)

            self.assertEqual(create_proposal_response.verdict, "ok")
            self.assertEqual(list_proposal_response.result["count"], 1)
            self.assertEqual(update_proposal_response.result["proposal"]["status"], "approved")
            self.assertEqual(proposal_approval_response.verdict, "ok")
            self.assertIsNotNone(proposal_approval_response.result["approval_request"]["id"])

            self.assertEqual(create_steering_response.verdict, "ok")
            self.assertEqual(list_steering_response.result["count"], 1)
            self.assertEqual(update_steering_response.result["approval"]["status"], "approved")
            self.assertEqual(steering_approval_response.verdict, "ok")
            self.assertIsNotNone(steering_approval_response.result["approval_request"]["id"])
            self.assertEqual(governance_audit_response.verdict, "ok")
            self.assertEqual(governance_audit_response.result["summary"]["mismatch_count"], 0)
            self.assertEqual(governance_repair_response.verdict, "ok")
            self.assertEqual(governance_repair_response.result["audit"]["status"], "ok")
            self.assertEqual(governance_repair_response.result["record"]["status"], "approval_requested")
            self.assertEqual(governance_review_queue_response.verdict, "ok")
            self.assertGreaterEqual(governance_review_queue_response.result["summary"]["actionable_count"], 1)
            self.assertGreaterEqual(len(governance_review_queue_response.result["recommendations"]), 1)

            self.assertEqual(document_summary_response.verdict, "ok")
            self.assertEqual(
                document_summary_response.result["summary"]["headline"],
                "Checklist for workflow approval and validation.",
            )
            self.assertEqual(add_document_link_response.verdict, "ok")
            self.assertEqual(list_document_links_response.result["count"], 1)
            self.assertEqual(delete_document_link_response.result["deleted"], True)

            self.assertEqual(create_namespace_response.verdict, "ok")
            self.assertEqual(list_namespaces_response.result["count"], 1)
            self.assertEqual(add_namespace_memory_response.verdict, "ok")
            self.assertEqual(namespace_memories_response.result["count"], 1)
            self.assertEqual(remove_namespace_memory_response.result["deleted"], True)

            self.assertEqual(documents_response.verdict, "ok")
            self.assertEqual(documents_response.result["count"], 1)
            self.assertEqual(documents_response.result["items"][0]["id"], "document-nav")

        asyncio.run(scenario())

    def test_governance_repair_tool_can_link_orphaned_approval(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-repair",
                    name="Workflow Repair",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-repair",
                {"title": "Repair proposal", "summary": "Needs approval rebind."},
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            orphaned_approval = await context.conversation_approval_repo.create(
                ApprovalRequest(
                    approval_type=ApprovalType.WORKFLOW_UPDATE,
                    status=ApprovalStatus.PENDING,
                    target_type=ApprovalTargetType.WORKFLOW,
                    target_id="workflow-repair",
                    requested_by_agent_id="main-agent",
                    requested_by_profile_id="profile-contract",
                    conversation_id="conversation-repair",
                    origin_message_id="message-repair",
                    summary="Orphaned workflow approval.",
                    metadata={"source": "workflow_service"},
                )
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            repair_response = await executor.run_async(
                "agency.workflow.governance.repair",
                {
                    "workflow_id": "workflow-repair",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                    "action": "link_approval_request",
                    "approval_request_id": orphaned_approval.id,
                },
                actor="user-contract",
            )

            refreshed_approval = await context.conversation_approval_repo.get(orphaned_approval.id)

            self.assertEqual(repair_response.verdict, "ok")
            self.assertEqual(repair_response.result["audit"]["status"], "ok")
            self.assertEqual(repair_response.result["record"]["approval_request_id"], orphaned_approval.id)
            self.assertEqual(repair_response.result["record"]["status"], "approval_requested")
            assert refreshed_approval is not None
            self.assertEqual(refreshed_approval.metadata.get("workflow_id"), "workflow-repair")
            self.assertEqual(refreshed_approval.metadata.get("proposal_id"), proposal_id)

        asyncio.run(scenario())

    def test_governance_remediation_tool_repairs_deterministic_drift(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-remediate",
                    name="Workflow Remediate",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-remediate",
                {
                    "title": "Status drift",
                    "summary": "Proposal status should match approval.",
                    "approval_request_id": "approval-missing",
                },
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            await WorkflowService(context).create_workflow_steering_approval(
                "workflow-remediate",
                {
                    "approval_id": "steering-1",
                    "recommended_action": "request_replan",
                    "reason": "Needs adoption.",
                    "approval_request_id": "approval-stale",
                },
            )
            proposal_approval = await context.conversation_approval_repo.create(
                ApprovalRequest(
                    approval_type=ApprovalType.WORKFLOW_UPDATE,
                    status=ApprovalStatus.APPROVED,
                    target_type=ApprovalTargetType.WORKFLOW,
                    target_id="workflow-remediate",
                    requested_by_agent_id="main-agent",
                    requested_by_profile_id="profile-contract",
                    conversation_id="conversation-remediate-1",
                    origin_message_id="message-remediate-1",
                    summary="Proposal approval.",
                    approved_by_user_id="user-contract",
                    metadata={"source": "workflow_service", "workflow_id": "workflow-remediate", "proposal_id": proposal_id},
                )
            )
            steering_approval = await context.conversation_approval_repo.create(
                ApprovalRequest(
                    approval_type=ApprovalType.WORKFLOW_UPDATE,
                    status=ApprovalStatus.PENDING,
                    target_type=ApprovalTargetType.WORKFLOW,
                    target_id="workflow-remediate",
                    requested_by_agent_id="main-agent",
                    requested_by_profile_id="profile-contract",
                    conversation_id="conversation-remediate-2",
                    origin_message_id="message-remediate-2",
                    summary="Steering approval.",
                    metadata={"source": "workflow_service", "workflow_id": "workflow-remediate", "approval_id": "steering-1"},
                )
            )
            await WorkflowService(context).update_workflow_improvement_proposal(
                "workflow-remediate",
                proposal_id,
                {"approval_request_id": proposal_approval.id, "status": "draft"},
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            dry_run_response = await executor.run_async(
                "agency.workflow.governance.remediate",
                {"workflow_id": "workflow-remediate", "dry_run": True},
                actor="user-contract",
            )
            apply_response = await executor.run_async(
                "agency.workflow.governance.remediate",
                {"workflow_id": "workflow-remediate"},
                actor="user-contract",
            )

            self.assertEqual(dry_run_response.verdict, "ok")
            self.assertEqual(dry_run_response.result["summary"]["planned_action_count"], 3)
            self.assertEqual(dry_run_response.result["summary"]["applied_action_count"], 0)
            self.assertEqual(apply_response.verdict, "ok")
            self.assertEqual(apply_response.result["summary"]["planned_action_count"], 3)
            self.assertEqual(apply_response.result["summary"]["applied_action_count"], 3)
            self.assertEqual(apply_response.result["audit_after"]["summary"]["mismatch_count"], 0)
            steering_records = apply_response.result["audit_after"]["steering_approvals"]
            proposal_records = apply_response.result["audit_after"]["proposals"]
            self.assertEqual(proposal_records[0]["record"]["status"], "approved")
            self.assertEqual(steering_records[0]["record"]["approval_request_id"], steering_approval.id)

        asyncio.run(scenario())

    def test_governance_review_queue_surfaces_evidence_and_repair_actions(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-review",
                    name="Workflow Review",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-review",
                {"title": "Needs evidence", "summary": "Proposal without supporting document."},
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            await WorkflowService(context).create_workflow_steering_approval(
                "workflow-review",
                {
                    "approval_id": "steering-review",
                    "recommended_action": "redirect_subagent",
                    "reason": "Steering packet missing approval link.",
                    "approval_request_id": "missing-approval",
                },
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            queue_response = await executor.run_async(
                "agency.workflow.governance.review-queue",
                {"workflow_id": "workflow-review"},
                actor="user-contract",
            )

            self.assertEqual(queue_response.verdict, "ok")
            self.assertEqual(queue_response.result["summary"]["proposal_count"], 1)
            self.assertEqual(queue_response.result["summary"]["steering_approval_count"], 1)
            self.assertEqual(queue_response.result["summary"]["remediation_candidate_count"], 1)
            self.assertGreaterEqual(queue_response.result["summary"]["actionable_count"], 2)
            self.assertEqual(queue_response.result["proposals"][0]["record_id"], proposal_id)
            self.assertEqual(queue_response.result["proposals"][0]["priority"], "approval")
            self.assertIn("attach_evidence", queue_response.result["proposals"][0]["next_actions"])
            self.assertEqual(queue_response.result["steering_approvals"][0]["priority"], "repair")
            recommendation_actions = {item["action"] for item in queue_response.result["recommendations"]}
            self.assertIn("preview_or_apply_remediation", recommendation_actions)
            self.assertIn("attach_workflow_evidence", recommendation_actions)

        asyncio.run(scenario())

    def test_governance_action_tool_executes_queue_actions(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-action",
                    name="Workflow Action",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            await context.uploaded_document_repo.create(
                UploadedDocument(
                    id="document-action",
                    filename="action.txt",
                    extracted_text="Evidence for governance action.",
                    content_sha256="action123",
                    text_characters=20,
                    estimated_tokens=5,
                    upload_mode=DocumentUploadMode.VECTOR,
                    scope="workflow",
                    workflow_id="workflow-action",
                    created_by_user_id="user-contract",
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-action",
                {"title": "Action proposal", "summary": "Needs approval and evidence."},
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            await WorkflowService(context).create_workflow_steering_approval(
                "workflow-action",
                {
                    "approval_id": "steering-action",
                    "recommended_action": "request_replan",
                    "reason": "Repair this stale approval reference.",
                    "approval_request_id": "missing-action-approval",
                },
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            approval_response = await executor.run_async(
                "agency.workflow.governance.act",
                {
                    "workflow_id": "workflow-action",
                    "action": "request_approval",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                },
                actor="user-contract",
            )
            evidence_response = await executor.run_async(
                "agency.workflow.governance.act",
                {
                    "workflow_id": "workflow-action",
                    "action": "attach_evidence",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                    "document_id": "document-action",
                    "summary": "Evidence attached from action tool.",
                },
                actor="user-contract",
            )
            remediation_response = await executor.run_async(
                "agency.workflow.governance.act",
                {
                    "workflow_id": "workflow-action",
                    "action": "apply_remediation",
                },
                actor="user-contract",
            )

            self.assertEqual(approval_response.verdict, "ok")
            self.assertEqual(approval_response.result["action"], "request_approval")
            self.assertIsNotNone(approval_response.result["result"]["approval_request"]["id"])
            self.assertEqual(evidence_response.verdict, "ok")
            self.assertEqual(evidence_response.result["action"], "attach_evidence")
            self.assertEqual(evidence_response.result["result"]["link"]["document_id"], "document-action")
            self.assertEqual(remediation_response.verdict, "ok")
            self.assertEqual(remediation_response.result["action"], "apply_remediation")
            self.assertGreaterEqual(remediation_response.result["result"]["summary"]["planned_action_count"], 1)

        asyncio.run(scenario())

    def test_governance_document_suggest_tool_ranks_workflow_evidence(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-doc-suggest",
                    name="Workflow Doc Suggest",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-doc-suggest",
                {
                    "title": "Validation checklist upgrade",
                    "summary": "Improve approval checklist and validation plan.",
                    "validation_plan": "Checklist validation for workflow approval.",
                    "tags": ["approval", "checklist"],
                },
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            await context.uploaded_document_repo.create(
                UploadedDocument(
                    id="document-match",
                    filename="approval-checklist.txt",
                    extracted_text="Workflow approval checklist and validation guidance for operators.",
                    content_sha256="match123",
                    text_characters=40,
                    estimated_tokens=10,
                    upload_mode=DocumentUploadMode.VECTOR,
                    scope="workflow",
                    workflow_id="workflow-doc-suggest",
                    created_by_user_id="user-contract",
                    metadata={
                        "upload_intelligence": {
                            "summary": "Checklist for workflow approval validation.",
                            "document_kind": "policy_sop",
                            "recommended": {"tags": ["approval", "checklist"]},
                        }
                    },
                )
            )
            await context.uploaded_document_repo.create(
                UploadedDocument(
                    id="document-weak",
                    filename="random-notes.txt",
                    extracted_text="Unrelated brainstorming notes.",
                    content_sha256="weak123",
                    text_characters=20,
                    estimated_tokens=5,
                    upload_mode=DocumentUploadMode.VECTOR,
                    scope="workflow",
                    workflow_id="workflow-doc-suggest",
                    created_by_user_id="user-contract",
                )
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            response = await executor.run_async(
                "agency.workflow.governance.document-suggest",
                {
                    "workflow_id": "workflow-doc-suggest",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                },
                actor="user-contract",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["count"], 1)
            self.assertEqual(response.result["items"][0]["document"]["id"], "document-match")
            self.assertGreater(response.result["items"][0]["score"], 0)
            self.assertIn("checklist", " ".join(response.result["items"][0]["matched_terms"]))

        asyncio.run(scenario())

    def test_governance_bundle_tool_plans_and_executes_attach_then_approval(self) -> None:
        async def scenario():
            context = create_test_api_context()
            await context.user_repo.create(
                UserDefinition(id="user-contract", email="contract@example.com", display_name="Contract User")
            )
            await context.workflow_repo.create(
                WorkflowDefinition(
                    id="workflow-bundle",
                    name="Workflow Bundle",
                    entrypoint="node-1",
                    metadata={"visible_to_main_agent": True},
                )
            )
            proposal_payload = await WorkflowService(context).create_workflow_improvement_proposal(
                "workflow-bundle",
                {
                    "title": "Approval checklist update",
                    "summary": "Attach evidence and request approval.",
                    "validation_plan": "Checklist validation and approval review.",
                    "tags": ["approval", "checklist"],
                },
            )
            proposal_id = str(proposal_payload["proposal"]["id"])
            await context.uploaded_document_repo.create(
                UploadedDocument(
                    id="document-bundle",
                    filename="bundle-checklist.txt",
                    extracted_text="Approval checklist and validation guide for workflow bundle review.",
                    content_sha256="bundle123",
                    text_characters=45,
                    estimated_tokens=12,
                    upload_mode=DocumentUploadMode.VECTOR,
                    scope="workflow",
                    workflow_id="workflow-bundle",
                    created_by_user_id="user-contract",
                    metadata={
                        "upload_intelligence": {
                            "summary": "Workflow approval checklist.",
                            "document_kind": "policy_sop",
                            "recommended": {"tags": ["approval", "checklist"]},
                        }
                    },
                )
            )
            registry = ToolRegistry(load_contracts())
            executor = ToolRuntimeExecutor(
                registry=registry,
                policy_engine=PolicyEngine(),
                tool_run_store=JsonlToolRunStore(Path(tempfile.mkdtemp()) / "tool-runs.jsonl"),
                context=context,
            )

            dry_run_response = await executor.run_async(
                "agency.workflow.governance.bundle",
                {
                    "workflow_id": "workflow-bundle",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                    "dry_run": True,
                },
                actor="user-contract",
            )
            apply_response = await executor.run_async(
                "agency.workflow.governance.bundle",
                {
                    "workflow_id": "workflow-bundle",
                    "record_kind": "improvement_proposal",
                    "record_id": proposal_id,
                },
                actor="user-contract",
            )

            self.assertEqual(dry_run_response.verdict, "ok")
            self.assertEqual(dry_run_response.result["suggestions"]["count"], 1)
            self.assertEqual(len(dry_run_response.result["planned_steps"]), 2)
            self.assertEqual(len(dry_run_response.result["applied_steps"]), 0)
            self.assertEqual(apply_response.verdict, "ok")
            self.assertEqual(apply_response.result["suggestions"]["items"][0]["document"]["id"], "document-bundle")
            self.assertEqual(len(apply_response.result["applied_steps"]), 2)
            self.assertEqual(apply_response.result["applied_steps"][0]["action"], "attach_top_suggestion")
            self.assertEqual(apply_response.result["applied_steps"][1]["action"], "request_approval")

        asyncio.run(scenario())

    def test_graph_context_tool_discovery_respects_feature_flag(self):
        previous = os.environ.get("AGENCY_GRAPH_CONTEXT_TOOLS_ENABLED")
        try:
            os.environ["AGENCY_GRAPH_CONTEXT_TOOLS_ENABLED"] = "false"
            reset_settings_cache()
            get_default_contract_registry.cache_clear()

            tool_ids = {tool.id for tool in list_builtin_tool_definitions()}
            contracts = {contract.name for contract in get_default_contract_registry().list_contracts()}

            self.assertNotIn("agency.graph.context", tool_ids)
            self.assertNotIn("agency.graph.context", contracts)
            self.assertNotIn("agency.graph.search", tool_ids)
            self.assertNotIn("agency.graph.search", contracts)
            self.assertNotIn("agency.graph.expand", tool_ids)
            self.assertNotIn("agency.graph.expand", contracts)
            self.assertNotIn("agency.graph.neighbors", tool_ids)
            self.assertNotIn("agency.graph.neighbors", contracts)
            self.assertNotIn("agency.graph.path", tool_ids)
            self.assertNotIn("agency.graph.path", contracts)
            self.assertNotIn("agency.graph.summarize-subgraph", tool_ids)
            self.assertNotIn("agency.graph.summarize-subgraph", contracts)
            self.assertNotIn("agency.graph.working-set.create", tool_ids)
            self.assertNotIn("agency.graph.working-set.create", contracts)
            self.assertNotIn("agency.graph.working-set.add", tool_ids)
            self.assertNotIn("agency.graph.working-set.add", contracts)
            self.assertNotIn("agency.graph.working-set.remove", tool_ids)
            self.assertNotIn("agency.graph.working-set.remove", contracts)
            self.assertNotIn("agency.graph.working-set.summarize", tool_ids)
            self.assertNotIn("agency.graph.working-set.summarize", contracts)
            self.assertNotIn("agency.graph.working-set.clear", tool_ids)
            self.assertNotIn("agency.graph.working-set.clear", contracts)
            self.assertNotIn("agency.graph.working-set.persist-context-pack", tool_ids)
            self.assertNotIn("agency.graph.working-set.persist-context-pack", contracts)
        finally:
            if previous is None:
                os.environ.pop("AGENCY_GRAPH_CONTEXT_TOOLS_ENABLED", None)
            else:
                os.environ["AGENCY_GRAPH_CONTEXT_TOOLS_ENABLED"] = previous
            reset_settings_cache()
            get_default_contract_registry.cache_clear()

    def test_main_agent_graph_context_policy_assigns_graph_tools_and_agent_config(self):
        async def scenario():
            context = create_test_api_context()
            await context.model_profile_repo.save(
                ModelProfileDefinition(id="profile-graph", name="Graph Model", provider="fake", model="fake-model")
            )
            profile = await MainAgentSetupService(context).create_main_agent(
                MainAgentSetupConfig(
                    agent_name="Graph Main Agent",
                    agent_description="Configured with graph context.",
                    agent_instructions="Use graph context when useful.",
                    model_profile_id="profile-graph",
                    profile_id="main-agent-graph-profile",
                    agent_id="main-agent-graph",
                    workflow_id="main-workflow-graph",
                    policy={
                        "can_answer_directly": True,
                        "can_trigger_workflows": True,
                        "can_manage_tools": True,
                        "can_manage_agents": True,
                        "can_manage_integrations": True,
                        "can_manage_memory": True,
                        "can_inspect_executions": True,
                        "can_run_commands": True,
                        "can_read_graph_context": True,
                        "graph_context_auto_retrieval_enabled": True,
                        "graph_context_subagent_steering_enabled": True,
                        "graph_context_coding_agent_resume_enabled": False,
                        "graph_context_default_intent": "handoff",
                        "graph_context_default_budget": "brief",
                    },
                )
            )
            agent = await context.agent_repo.get(profile.agent_id)
            assert agent is not None
            return agent

        agent = asyncio.run(scenario())

        self.assertIn("agency.graph.context", agent.tool_ids)
        self.assertIn("agency.graph.search", agent.tool_ids)
        self.assertTrue(agent.graph_context.enabled)
        self.assertTrue(agent.graph_context.auto_retrieval_enabled)
        self.assertTrue(agent.graph_context.subagent_steering_enabled)
        self.assertFalse(agent.graph_context.coding_agent_resume_enabled)
        self.assertEqual(agent.graph_context.default_intent, "handoff")
        self.assertEqual(agent.graph_context.default_budget, "brief")

    def test_main_agent_default_policy_enables_balanced_graph_context(self):
        async def scenario():
            context = create_test_api_context()
            await context.model_profile_repo.save(
                ModelProfileDefinition(id="profile-main-default", name="Main Default", provider="fake", model="fake-model")
            )
            profile = await MainAgentSetupService(context).create_main_agent(
                MainAgentSetupConfig(
                    agent_name="Default Graph Main Agent",
                    agent_description="Configured with default graph context.",
                    agent_instructions="Use graph context when useful.",
                    model_profile_id="profile-main-default",
                    profile_id="main-agent-default-graph-profile",
                    agent_id="main-agent-default-graph",
                    workflow_id="main-workflow-default-graph",
                )
            )
            agent = await context.agent_repo.get(profile.agent_id)
            assert agent is not None
            return agent

        agent = asyncio.run(scenario())

        self.assertIn("agency.graph.context", agent.tool_ids)
        self.assertTrue(agent.graph_context.enabled)
        self.assertTrue(agent.graph_context.auto_retrieval_enabled)
        self.assertTrue(agent.graph_context.subagent_steering_enabled)
        self.assertFalse(agent.graph_context.coding_agent_resume_enabled)
        self.assertEqual(agent.graph_context.default_intent, "handoff")
        self.assertEqual(agent.graph_context.default_budget, "balanced")

    def test_policy_denies_unallowlisted_repos_dangerous_paths_and_high_risk_secrets(self):
        verdict = PolicyEngine(allowed_repos=[]).evaluate(
            "sandbox-edit",
            {
                "repo": "/tmp/not-allowed",
                "ref": "main",
                "changes": [
                    {"path": "src/.env.local", "patch": "+OPENAI_API_KEY=sk-test"},
                    {"path": ".git/config", "patch": "+[remote]\n"},
                ],
            },
        )

        self.assertEqual(verdict.outcome, "deny")
        rule_ids = {rule.id for rule in verdict.rules if rule.outcome == "deny"}
        self.assertIn("repo-allowlist", rule_ids)
        self.assertIn("no-dangerous-paths", rule_ids)
        self.assertIn("no-secrets", rule_ids)

    def test_runtime_returns_signed_structured_dry_run_response(self):
        with tempfile.TemporaryDirectory() as tmp:
            repo = Path(tmp)
            subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True, text=True)
            (repo / "README.md").write_text("hello\n", encoding="utf-8")

            store = JsonlToolRunStore(repo / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(policy_engine=PolicyEngine(allowed_repos=[str(repo)]), run_store=store)
            response = executor.run(
                "sandbox-edit",
                {
                    "repo": str(repo),
                    "ref": "main",
                    "changes": [{"path": "README.md", "patch": README_PATCH}],
                    "dryRun": True,
                },
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.errors, [])
            self.assertTrue(response.signature.startswith("sha256:"))
            self.assertEqual(response.filesChanged[0].path, "README.md")
            self.assertEqual(response.patch, README_PATCH)
            self.assertEqual((repo / "README.md").read_text(encoding="utf-8"), "hello\n")
            records = store.list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].tool_name, "sandbox-edit")
            self.assertEqual(records[0].tool_version, "1.0")
            self.assertEqual(records[0].actor, "user-runtime")
            self.assertEqual(records[0].verdict, "ok")
            self.assertIn("filesystem", records[0].risk_labels)
            self.assertIn("local_privileged_execution", records[0].risk_labels)
            self.assertEqual(records[0].risk_metadata_json["localPrivilegedExecution"], True)
            self.assertTrue(records[0].input_hash.startswith("sha256:"))
            self.assertTrue(records[0].output_hash.startswith("sha256:"))

    def test_runtime_runs_contract_backed_tool_list(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(run_store=store)
            response = executor.run("agency.tool.list", {}, actor="user-runtime")

            self.assertEqual(response.verdict, "ok")
            self.assertIsNotNone(response.policyVerdict)
            self.assertIsNotNone(response.result)
            assert response.result is not None
            self.assertEqual(response.result["count"], len(list_builtin_tool_definitions()))
            tool_ids = {item["id"] for item in response.result["items"]}
            self.assertIn("agency.speech.listen", tool_ids)
            self.assertIn("agency.tool.list", tool_ids)
            self.assertIn("agency.graph.context", tool_ids)
            self.assertIn("agency.graph.search", tool_ids)
            self.assertIn("agency.graph.expand", tool_ids)
            self.assertIn("agency.graph.neighbors", tool_ids)
            self.assertIn("agency.graph.path", tool_ids)
            self.assertIn("agency.graph.summarize-subgraph", tool_ids)
            self.assertIn("agency.command.run", tool_ids)
            records = store.list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].tool_name, "agency.tool.list")

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_runs_contract_backed_http_request(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                run_store=store,
            )
            response = executor.run(
                "agency.http.request",
                {"url": "https://api.example.test/items", "method": "GET", "verify_ssl": True},
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status_code"], 200)
            self.assertEqual(response.result["response"], {"ok": True})
            self.assertEqual(response.result["method"], "GET")
            self.assertTrue(response.signature.startswith("sha256:"))
            self.assertEqual(store.list_records()[0].tool_name, "agency.http.request")
            mock_execute.assert_called_once()

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_routes_http_request_through_onecli_proxy(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        "ONECLI_GATEWAY_CA_BUNDLE_PATH": "/tmp/onecli-ca.pem",
                        "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_AGENT_TOKEN",
                        "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                        "ONECLI_AGENT_TOKEN": "test-onecli-agent-token",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = executor.run(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "verify_ssl": True,
                        "credential_mode": "onecli",
                    },
                    actor="user-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["credential_mode"], "onecli")
            self.assertEqual(response.result["onecli"]["gateway_url"], "http://onecli.local:10255")
            self.assertEqual(response.result["onecli"]["gateway_mode"], "proxy")
            self.assertEqual(response.result["onecli"]["target_scheme"], "https")
            self.assertEqual(response.result["onecli"]["target_host"], "api.example.test")
            self.assertIsNone(response.result["onecli"]["target_port"])
            self.assertTrue(response.result["onecli"]["correlation_id"].startswith("onecli-http:"))
            self.assertTrue(response.result["onecli"]["agent_token_secret_ref_configured"])
            self.assertEqual(response.result["onecli"]["agent_identity"]["mapping"], "development_global_fallback")
            self.assertEqual(response.result["onecli"]["agent_identity"]["agency_actor"], "user-runtime")
            self.assertTrue(response.result["onecli"]["agent_identity"]["agent_token_secret_ref_configured"])
            self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(response.result))
            mock_execute.assert_called_once()
            kwargs = mock_execute.call_args.kwargs
            self.assertEqual(
                kwargs["proxies"],
                {
                    "http": "http://x:test-onecli-agent-token@onecli.local:10255",
                    "https": "http://x:test-onecli-agent-token@onecli.local:10255",
                },
            )
            self.assertEqual(kwargs["ca_bundle_path"], "/tmp/onecli-ca.pem")
            self.assertEqual(kwargs["headers"]["X-Agency-OneCLI-Correlation-ID"], response.result["onecli"]["correlation_id"])
            self.assertEqual(kwargs["headers"]["X-Agency-User-ID"], "user-runtime")
            self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(kwargs["headers"]))

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_prefers_user_onecli_identity_mapping_for_proxy_auth(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}

        async def run_assertions():
            context = create_test_api_context()
            await context.onecli_identity_mapping_repo.create(
                OneCLIIdentityMapping(
                    id="mapping-user-runtime",
                    owner_user_id="user-runtime",
                    name="Runtime User OneCLI",
                    onecli_agent_id="onecli-agent-user-runtime",
                    agent_token_secret_ref="env://ONECLI_USER_RUNTIME_TOKEN",
                )
            )
            with tempfile.TemporaryDirectory() as tmp, patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_GLOBAL_TOKEN",
                        "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                        "ONECLI_GLOBAL_TOKEN": "global-onecli-token",
                        "ONECLI_USER_RUNTIME_TOKEN": "mapped-onecli-token",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    context=context,
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = await executor.run_async(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "verify_ssl": True,
                        "credential_mode": "onecli",
                    },
                    actor="user-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["credential_mode"], "onecli")
            self.assertTrue(response.result["onecli"]["agent_token_secret_ref_configured"])
            self.assertNotIn("ONECLI_USER_RUNTIME_TOKEN", str(response.result))
            self.assertNotIn("mapped-onecli-token", str(response.result))
            kwargs = mock_execute.call_args.kwargs
            self.assertEqual(
                kwargs["proxies"],
                {
                    "http": "http://x:mapped-onecli-token@onecli.local:10255",
                    "https": "http://x:mapped-onecli-token@onecli.local:10255",
                },
            )
            self.assertEqual(kwargs["headers"]["X-Agency-OneCLI-Correlation-ID"], response.result["onecli"]["correlation_id"])
            self.assertEqual(kwargs["headers"]["X-Agency-User-ID"], "user-runtime")
            self.assertNotIn("global-onecli-token", str(kwargs))
            actions = context.runtime_operations.snapshot_dict()["recent_actions"]
            used_actions = [
                item for item in actions
                if item["action"] == "onecli.identity_mapping.used"
            ]
            self.assertTrue(used_actions)
            self.assertEqual(used_actions[-1]["mapping_id"], "mapping-user-runtime")
            self.assertEqual(used_actions[-1]["owner_user_id"], "user-runtime")
            self.assertTrue(used_actions[-1]["agent_token_secret_ref_configured"])
            self.assertNotIn("ONECLI_USER_RUNTIME_TOKEN", str(used_actions))
            self.assertNotIn("mapped-onecli-token", str(used_actions))

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_does_not_use_another_users_onecli_identity_mapping(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}

        async def run_assertions():
            context = create_test_api_context()
            await context.onecli_identity_mapping_repo.create(
                OneCLIIdentityMapping(
                    id="mapping-owner-runtime",
                    owner_user_id="user-owner-runtime",
                    name="Owner Runtime OneCLI",
                    onecli_agent_id="onecli-agent-owner-runtime",
                    agent_token_secret_ref="env://ONECLI_OWNER_RUNTIME_TOKEN",
                )
            )
            with tempfile.TemporaryDirectory() as tmp, patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_GLOBAL_TOKEN",
                        "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                        "ONECLI_GLOBAL_TOKEN": "global-onecli-token",
                        "ONECLI_OWNER_RUNTIME_TOKEN": "owner-mapped-onecli-token",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    context=context,
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = await executor.run_async(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "verify_ssl": True,
                        "credential_mode": "onecli",
                    },
                    actor="user-other-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "ok")
            self.assertNotIn("ONECLI_OWNER_RUNTIME_TOKEN", str(response.result))
            self.assertNotIn("owner-mapped-onecli-token", str(response.result))
            kwargs = mock_execute.call_args.kwargs
            self.assertEqual(
                kwargs["proxies"],
                {
                    "http": "http://x:global-onecli-token@onecli.local:10255",
                    "https": "http://x:global-onecli-token@onecli.local:10255",
                },
            )
            self.assertEqual(kwargs["headers"]["X-Agency-OneCLI-Correlation-ID"], response.result["onecli"]["correlation_id"])
            self.assertEqual(kwargs["headers"]["X-Agency-User-ID"], "user-other-runtime")
            self.assertNotIn("owner-mapped-onecli-token", str(kwargs))
            actions = context.runtime_operations.snapshot_dict()["recent_actions"]
            used_actions = [
                item for item in actions
                if item["action"] == "onecli.identity_mapping.used"
            ]
            self.assertEqual(used_actions, [])
            fallback_actions = [
                item for item in actions
                if item["action"] == "onecli.global_agent_token_fallback.used"
            ]
            self.assertTrue(fallback_actions)
            self.assertEqual(fallback_actions[-1]["owner_user_id"], "user-other-runtime")
            self.assertTrue(fallback_actions[-1]["agent_token_secret_ref_configured"])
            self.assertNotIn("env://ONECLI_GLOBAL_TOKEN", str(fallback_actions))

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_global_onecli_fallback_is_disabled_by_default(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}

        async def run_assertions():
            context = create_test_api_context()
            with tempfile.TemporaryDirectory() as tmp, patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_GLOBAL_TOKEN",
                        "ONECLI_GLOBAL_TOKEN": "global-onecli-token",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    context=context,
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = await executor.run_async(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "verify_ssl": True,
                        "credential_mode": "onecli",
                    },
                    actor="user-without-mapping",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["onecli"]["agent_identity"]["mapping"], "none")
            kwargs = mock_execute.call_args.kwargs
            self.assertEqual(
                kwargs["proxies"],
                {
                    "http": "http://onecli.local:10255",
                    "https": "http://onecli.local:10255",
                },
            )
            self.assertNotIn("global-onecli-token", str(kwargs))
            actions = context.runtime_operations.snapshot_dict()["recent_actions"]
            fallback_actions = [
                item for item in actions
                if item["action"] == "onecli.global_agent_token_fallback.used"
            ]
            self.assertEqual(fallback_actions, [])

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_denies_onecli_http_request_when_onecli_disabled(self, mock_execute):
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(os.environ, {"ONECLI_ENABLED": "false"}, clear=False):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = executor.run(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "credential_mode": "onecli",
                    },
                    actor="user-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "deny")
            self.assertIn("ONECLI_ENABLED is false", response.errors[0])
            mock_execute.assert_not_called()

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_denies_onecli_http_request_when_global_kill_switch_enabled(self, mock_execute):
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_EXTERNAL_CALLS_DISABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = executor.run(
                    "agency.http.request",
                    {
                        "url": "https://api.example.test/items",
                        "method": "GET",
                        "credential_mode": "onecli",
                    },
                    actor="user-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "deny")
            self.assertIn("ONECLI_EXTERNAL_CALLS_DISABLED", response.errors[0])
            mock_execute.assert_not_called()

    @patch("app.core.outbound_http.socket.getaddrinfo")
    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_fails_closed_when_onecli_gateway_unavailable_in_production(self, mock_execute, getaddrinfo):
        getaddrinfo.return_value = [(2, 1, 6, "", ("93.184.216.34", 443))]
        mock_execute.side_effect = RuntimeError("onecli gateway unavailable")

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with tempfile.TemporaryDirectory() as tmp, patch.dict(
                        os.environ,
                        {
                            "APP_ENV": "production",
                            "DATABASE_URL": "sqlite+aiosqlite:///:memory:",
                            "AGENCY_INTERNAL_API_KEY": "trusted-key",
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                            "TOOL_HTTP_ALLOWED_HOSTS": "api.example.test",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    executor = ToolRuntimeExecutor(
                        policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                        run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                    )
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                        },
                        actor="user-runtime",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(response.verdict, "deny")
                    self.assertIsNone(response.result)
                    self.assertIn("onecli gateway unavailable", response.errors[0])
                    kwargs = mock_execute.call_args.kwargs
                    self.assertEqual(kwargs["proxies"]["https"], "http://onecli.local:10255")

                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    self.assertEqual(
                        [event.metadata.get("semanticType") for event in onecli_events],
                        ["onecli.http.request.started", "onecli.http.request.failed"],
                    )
                    self.assertEqual(onecli_events[-1].metadata["verdict"], "deny")
                    self.assertTrue(onecli_events[-1].metadata["fail_closed"])
                    self.assertEqual(onecli_events[-1].level.value, "error")
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_defaults_http_request_to_onecli_when_forced(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_FORCE_FOR_HTTP_TOOLS": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                executor = ToolRuntimeExecutor(
                    policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                response = executor.run(
                    "agency.http.request",
                    {"url": "https://api.example.test/items", "method": "GET"},
                    actor="user-runtime",
                )
            reset_settings_cache()

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["credential_mode"], "onecli")
            self.assertEqual(mock_execute.call_args.kwargs["proxies"]["https"], "http://onecli.local:10255")

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_denies_unapproved_mutation_before_onecli_proxy(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with tempfile.TemporaryDirectory() as tmp, patch.dict(
                        os.environ,
                        {
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    executor = ToolRuntimeExecutor(
                        policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                        run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                    )
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "POST",
                            "body": {"name": "created-by-onecli-approval-test"},
                            "credential_mode": "onecli",
                        },
                        actor="user-runtime",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(response.verdict, "deny")
                    rules_by_id = {rule.id: rule for rule in response.policyVerdict.rules}
                    self.assertEqual(rules_by_id["http-mutation-approval-context"].outcome, "deny")
                    self.assertIn("requires an explicitly approved", rules_by_id["http-mutation-approval-context"].reason)
                    mock_execute.assert_not_called()

                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    semantic_types = [event.metadata.get("semanticType") for event in events]
                    policy_index = semantic_types.index("tool.policy.completed")
                    self.assertNotIn("onecli.http.request.started", semantic_types)
                    self.assertEqual(events[policy_index].metadata["verdict"], "deny")
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_denies_onecli_http_request_with_direct_auth_header(self, mock_execute):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
            )
            for header_name in sorted(ONECLI_BLOCKED_HEADER_NAMES):
                with self.subTest(header_name=header_name):
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                            "headers": {header_name: "raw-token"},
                        },
                        actor="user-runtime",
                    )

                    self.assertEqual(response.verdict, "deny")
                    self.assertIn("credential-bearing headers", response.errors[0])
                    self.assertIn(header_name, response.errors[0])
            mock_execute.assert_not_called()

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_denies_onecli_http_request_with_direct_auth_query_param(self, mock_execute):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
            )
            for query_name in sorted(ONECLI_BLOCKED_QUERY_PARAM_NAMES):
                with self.subTest(query_name=query_name):
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                            "query_params": {query_name: "raw-key"},
                        },
                        actor="user-runtime",
                    )

                    self.assertEqual(response.verdict, "deny")
                    self.assertIn("credential-bearing query parameters", response.errors[0])
                    self.assertIn(query_name, response.errors[0])
            mock_execute.assert_not_called()

    def test_runtime_denies_contract_backed_http_request_disallowed_host(self):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
            )
            response = executor.run(
                "agency.http.request",
                {"url": "https://blocked.example.test/items", "method": "GET"},
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "deny")
            self.assertIsNone(response.result)
            self.assertIn("host is not allowlisted", response.errors[0])

    def test_runtime_runs_contract_backed_tool_get(self):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            response = executor.run("agency.tool.get", {"tool_id": "agency.http.request"}, actor="user-runtime")

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status"], "ok")
            self.assertEqual(response.result["tool"]["id"], "agency.http.request")

    def test_runtime_runs_context_backed_workflow_list_and_get(self):
        async def run_assertions():
            with tempfile.TemporaryDirectory() as tmp:
                context = create_test_api_context()
                workflow = await context.workflow_repo.create(_workflow_definition())
                executor = ToolRuntimeExecutor(
                    context=context,
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )

                list_response = await executor.run_async("agency.workflow.list", {}, actor="user-runtime")
                get_response = await executor.run_async(
                    "agency.workflow.get",
                    {"workflow_id": workflow.id},
                    actor="user-runtime",
                )

                self.assertEqual(list_response.verdict, "ok")
                self.assertEqual(list_response.result["status"], "ok")
                self.assertEqual(list_response.result["workflows"][0]["id"], workflow.id)
                self.assertEqual(get_response.verdict, "ok")
                self.assertEqual(get_response.result["workflow"]["id"], workflow.id)
                self.assertEqual(get_response.result["summary"]["input_keys"], ["topic"])

        asyncio.run(run_assertions())

    def test_runtime_runs_context_backed_memory_crud(self):
        async def run_assertions():
            with tempfile.TemporaryDirectory() as tmp:
                context = create_test_api_context()
                await context.user_repo.create(UserDefinition(id="user-memory", email="memory@example.com"))
                executor = ToolRuntimeExecutor(
                    context=context,
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )

                remember = await executor.run_async(
                    "agency.memory.remember",
                    {
                        "scope": "user",
                        "content": "Use concise contract updates.",
                        "summary": "Prefers concise updates.",
                        "tags": ["preference"],
                    },
                    actor="user-memory",
                )
                memory_id = remember.result["memory"]["id"]
                listed = await executor.run_async("agency.memory.list", {"scope": "user", "query": "concise"}, actor="user-memory")
                updated = await executor.run_async(
                    "agency.memory.update",
                    {"memory_id": memory_id, "summary": "Prefers concise engineering updates."},
                    actor="user-memory",
                )
                deleted = await executor.run_async("agency.memory.delete", {"memory_id": memory_id}, actor="user-memory")

                self.assertEqual(remember.verdict, "ok")
                self.assertEqual(remember.result["memory"]["created_by_user_id"], "user-memory")
                self.assertEqual(listed.result["memories"][0]["id"], memory_id)
                self.assertEqual(updated.result["memory"]["summary"], "Prefers concise engineering updates.")
                self.assertTrue(deleted.result["deleted"])

        asyncio.run(run_assertions())

    def test_runtime_runs_memory_catalog_exclusions_and_workflow_links(self):
        async def run_assertions():
            with tempfile.TemporaryDirectory() as tmp:
                context = create_test_api_context()
                await context.user_repo.create(UserDefinition(id="user-memory", email="memory@example.com"))
                await context.workflow_repo.create(
                    WorkflowDefinition(
                        id="workflow-memory-runtime",
                        name="Memory Runtime Workflow",
                        description="Workflow for memory link contract tests.",
                        entrypoint="task-memory",
                        agent_definitions=[AgentDefinition(id="agent-memory", name="Memory Agent")],
                        task_definitions=[
                            TaskDefinition(
                                id="task-memory",
                                name="Use memory",
                                description="Use linked memory.",
                                agent_id="agent-memory",
                            )
                        ],
                        metadata={"owner_ids": ["user-memory"], "created_by": "user-memory"},
                    )
                )
                executor = ToolRuntimeExecutor(
                    context=context,
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )

                remember = await executor.run_async(
                    "agency.memory.remember",
                    {
                        "scope": "user",
                        "content": "Use linked memory only when the workflow asks for it.",
                        "summary": "Linked memory preference.",
                        "tags": ["workflow-link"],
                    },
                    actor="user-memory",
                )
                memory_id = remember.result["memory"]["id"]

                catalog = await executor.run_async(
                    "agency.memory.catalog",
                    {"scope": "user", "query": "linked memory", "target_type": "task", "target_id": "task-memory"},
                    actor="user-memory",
                )
                exclusion = await executor.run_async(
                    "agency.memory.exclusions.add",
                    {
                        "memory_id": memory_id,
                        "target_type": "task",
                        "target_id": "task-memory",
                        "reason": "Only attach explicitly.",
                    },
                    actor="user-memory",
                )
                exclusions = await executor.run_async(
                    "agency.memory.exclusions.list",
                    {"memory_id": memory_id, "target_type": "task", "target_id": "task-memory"},
                    actor="user-memory",
                )
                deleted_exclusion = await executor.run_async(
                    "agency.memory.exclusions.delete",
                    {
                        "memory_id": memory_id,
                        "exclusion_id": exclusion.result["exclusion"]["id"],
                    },
                    actor="user-memory",
                )
                read_link = await executor.run_async(
                    "agency.workflow.memory-links.add",
                    {
                        "workflow_id": "workflow-memory-runtime",
                        "target_type": "task",
                        "target_id": "task-memory",
                        "ref_type": "memory",
                        "ref_id": memory_id,
                        "access_mode": "read",
                    },
                    actor="user-memory",
                )
                blocked_update = await executor.run_async(
                    "agency.memory.update",
                    {
                        "memory_id": memory_id,
                        "summary": "Should be blocked by read-only link.",
                        "workflow_id": "workflow-memory-runtime",
                        "target_type": "task",
                        "target_id": "task-memory",
                    },
                    actor="user-memory",
                )
                write_link = await executor.run_async(
                    "agency.workflow.memory-links.add",
                    {
                        "workflow_id": "workflow-memory-runtime",
                        "target_type": "task",
                        "target_id": "task-memory",
                        "ref_type": "memory",
                        "ref_id": memory_id,
                        "access_mode": "read_write",
                    },
                    actor="user-memory",
                )
                updated = await executor.run_async(
                    "agency.memory.update",
                    {
                        "memory_id": memory_id,
                        "summary": "Updated through read-write link.",
                        "workflow_id": "workflow-memory-runtime",
                        "target_type": "task",
                        "target_id": "task-memory",
                    },
                    actor="user-memory",
                )
                links = await executor.run_async(
                    "agency.workflow.memory-links.list",
                    {"workflow_id": "workflow-memory-runtime"},
                    actor="user-memory",
                )
                deleted_link = await executor.run_async(
                    "agency.workflow.memory-links.delete",
                    {"workflow_id": "workflow-memory-runtime", "link_id": write_link.result["link"]["id"]},
                    actor="user-memory",
                )

                self.assertEqual(catalog.verdict, "ok")
                self.assertTrue(
                    any(
                        item["id"] == memory_id
                        for group in catalog.result["catalog"]["groups"]
                        for item in group["items"]
                    )
                )
                self.assertEqual(exclusion.result["exclusion"]["targetType"], "task")
                self.assertEqual(exclusions.result["items"][0]["memoryId"], memory_id)
                self.assertTrue(deleted_exclusion.result["deleted"])
                self.assertEqual(read_link.result["link"]["accessMode"], "read")
                self.assertEqual(blocked_update.verdict, "warn")
                self.assertIn("read_write", blocked_update.result["error"])
                self.assertEqual(write_link.result["link"]["accessMode"], "read_write")
                self.assertEqual(updated.result["memory"]["summary"], "Updated through read-write link.")
                self.assertGreaterEqual(len(links.result["items"]), 2)
                self.assertTrue(deleted_link.result["deleted"])

        asyncio.run(run_assertions())

    def test_runtime_marks_direct_proposal_tools_as_conversation_context_required(self):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            response = executor.run(
                "agency.workflow.propose-create",
                {"goal": "Create a workflow that drafts a report."},
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "warn")
            self.assertEqual(response.result["status"], "requires_conversation_context")
            self.assertIn("requires conversation/profile/origin-message approval context", response.errors[0])

    @patch("app.tools.runtime.executor.open_browser")
    def test_runtime_runs_contract_backed_browser_open(self, mock_open_browser):
        mock_open_browser.return_value = {
            "session_id": "brs_runtime",
            "interactive": True,
            "engine": "playwright",
            "url": "https://example.test",
            "title": "Example",
            "runtime_root": "/tmp/browser-runtime",
            "message": "Browser started.",
        }
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(run_store=store)
            response = executor.run(
                "agency.browser.open",
                {"url": "https://example.test"},
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status"], "ok")
            self.assertEqual(response.result["output"]["title"], "Example")
            self.assertTrue(response.signature.startswith("sha256:"))
            records = store.list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].tool_name, "agency.browser.open")
            mock_open_browser.assert_called_once()
            _, call_kwargs = mock_open_browser.call_args
            self.assertEqual(call_kwargs["_browser_owner"], {"actor": "user-runtime"})

    @patch("app.tools.runtime.executor.click_element")
    def test_runtime_denies_contract_backed_browser_mutation_without_approval(self, mock_click):
        mock_click.return_value = "Clicked element matching instruction: Submit"
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            response = executor.run("agency.browser.click", {"instruction": "Submit"}, actor="user-runtime")

            self.assertEqual(response.verdict, "deny")
            self.assertIsNone(response.result)
            self.assertIn("browser mutation", response.policyVerdict.rules[1].reason)
            mock_click.assert_not_called()

    def test_runtime_routes_all_browser_tools_through_contract_executor(self):
        browser_cases = [
            ("agency.browser.screenshot", {}, "screenshot", "screenshot ok"),
            ("agency.browser.analyze-screenshot", {"text": "Analyze the page"}, "screenshot_and_analyse", "analysis ok"),
            (
                "agency.browser.extract-screenshot",
                {"text": "Extract content"},
                "screenshot_and_extract",
                {"page_type": "generic", "page_url": "https://example.test", "content": {"summary": "ok", "text": "ok"}},
            ),
            ("agency.browser.scroll", {"scroll_direction": "down 1"}, "scroll_page", "scroll ok"),
            ("agency.browser.select-option", {"instruction": "Select 'A'"}, "select_dropdown", "select ok"),
            ("agency.browser.type-text", {"instruction": "Type 'hello'"}, "send_keys", "type ok"),
            (
                "agency.browser.verify-content",
                {"text": "hello"},
                "verify_content",
                {"Verification Reasoning": "found", "Verification Score": 100, "Challenge Detected": False},
            ),
            ("agency.browser.close", {}, "terminate_browser", {"Success Message": "closed"}),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            for tool_name, payload, handler_name, raw_result in browser_cases:
                with self.subTest(tool_name=tool_name), patch(f"app.tools.runtime.executor.{handler_name}") as mock_handler:
                    mock_handler.return_value = raw_result
                    response = executor.run(tool_name, payload, actor="approved/user-runtime")

                    self.assertEqual(response.verdict, "ok")
                    self.assertEqual(response.result["status"], "ok")
                    self.assertEqual(response.result["output"], raw_result)

    @patch("app.tools.runtime.executor.request_human_input")
    def test_runtime_runs_contract_backed_human_ask(self, mock_request_human_input):
        mock_request_human_input.return_value = {"status": "received", "response": "Proceed"}
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            response = executor.run(
                "agency.human.ask",
                {"query": "Should I proceed?", "timeout_seconds": 1},
                actor="user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["response"], "Proceed")
            mock_request_human_input.assert_called_once_with("Should I proceed?", process_id=None, timeout=1)

    def test_runtime_requires_confirmation_for_sensitive_memory(self):
        async def run_assertions():
            with tempfile.TemporaryDirectory() as tmp:
                context = create_test_api_context()
                await context.user_repo.create(UserDefinition(id="user-memory", email="memory@example.com"))
                executor = ToolRuntimeExecutor(
                    context=context,
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )

                response = await executor.run_async(
                    "agency.memory.remember",
                    {
                        "scope": "user",
                        "content": "My password is example",
                        "sensitive": True,
                        "confirmed": False,
                    },
                    actor="user-memory",
                )

                self.assertEqual(response.verdict, "warn")
                self.assertEqual(response.result["status"], "error")
                self.assertIn("Sensitive memory writes require explicit user confirmation", response.errors[0])

        asyncio.run(run_assertions())

    def test_runtime_runs_contract_backed_command(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(run_store=store)
            response = executor.run(
                "agency.command.run",
                {"command": "printf 'contract-command\\n'", "mode": "bash", "timeout_seconds": 2},
                actor="approved/user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertIsNotNone(response.result)
            assert response.result is not None
            self.assertEqual(response.result["status"], "ok")
            self.assertEqual(response.result["exit_code"], 0)
            self.assertEqual(response.result["stdout"], "contract-command")
            self.assertTrue(response.signature.startswith("sha256:"))
            records = store.list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].tool_name, "agency.command.run")

    def test_runtime_denies_blocked_contract_backed_command(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(run_store=store)
            blocked_commands = {
                "git push origin main": "git push is blocked",
                "sed -n '1,5p' .env": "reading .env files is blocked",
                "awk '{print}' .env": "reading .env files is blocked",
                "python3 -c \"print(open('.env').read())\"": "reading .env files is blocked",
                "python3 -c \"print(open('/root/.ssh/id_rsa').read())\"": "reading SSH credentials is blocked",
            }

            for command, reason in blocked_commands.items():
                with self.subTest(command=command):
                    response = executor.run(
                        "agency.command.run",
                        {"command": command, "mode": "bash"},
                        actor="approved/user-runtime",
                    )

                    self.assertEqual(response.verdict, "deny")
                    self.assertIsNone(response.result)
                    self.assertTrue(response.errors)
                    self.assertIn(reason, response.errors[0])

            records = store.list_records()
            self.assertEqual(len(records), len(blocked_commands))
            self.assertTrue(all(record.verdict == "deny" for record in records))

    def test_runtime_runs_contract_backed_file_write_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = JsonlToolRunStore(root / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root)]),
                run_store=store,
            )
            response = executor.run(
                "agency.file.write-text",
                {
                    "base_folder": str(root),
                    "filename": "notes/out.txt",
                    "content": "contract file\n",
                    "mode": "write",
                },
                actor="approved/user-runtime",
            )

            written = root / "notes" / "out.txt"
            self.assertEqual(response.verdict, "ok")
            self.assertEqual(written.read_text(encoding="utf-8"), "contract file\n")
            self.assertIsNotNone(response.result)
            assert response.result is not None
            self.assertEqual(response.result["status"], "success")
            self.assertEqual(response.filesChanged[0].path, str(written.resolve()))
            records = store.list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].tool_name, "agency.file.write-text")

    def test_runtime_denies_contract_backed_file_write_outside_allowlist(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = JsonlToolRunStore(root / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root / "allowed")]),
                run_store=store,
            )
            response = executor.run(
                "agency.file.write-text",
                {
                    "base_folder": str(root / "blocked"),
                    "filename": "out.txt",
                    "content": "blocked",
                    "mode": "write",
                },
                actor="approved/user-runtime",
            )

            self.assertEqual(response.verdict, "deny")
            self.assertIsNone(response.result)
            self.assertFalse((root / "blocked" / "out.txt").exists())
            self.assertIn("not under an allowlisted directory", response.errors[0])

    @patch("app.tools.implementations.documents.upload_to_s3")
    def test_runtime_runs_contract_backed_markdown_to_word(self, mock_upload):
        mock_upload.return_value = {"uploaded_files": ["user_approved-user/workflow_reports/run_proc-1/report.docx"]}
        with tempfile.TemporaryDirectory() as tmp:
            store = JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl")
            executor = ToolRuntimeExecutor(run_store=store)
            response = executor.run(
                "agency.document.markdown-to-word",
                {
                    "markdown_text": "# Report\n\nBody",
                    "filename": "report.docx",
                    "img_directory": "reports",
                    "process_id": "proc-1",
                    "run_by": "approved-user",
                },
                actor="approved/user-runtime",
            )

            self.assertEqual(response.verdict, "ok")
            self.assertIsNotNone(response.result)
            assert response.result is not None
            self.assertEqual(response.result["status"], "success")
            self.assertEqual(response.result["storage_uri"], "s3://mybucket/user_approved-user/workflow_reports/run_proc-1/report.docx")
            self.assertTrue(response.signature.startswith("sha256:"))
            self.assertEqual(store.list_records()[0].tool_name, "agency.document.markdown-to-word")

    def test_runtime_denies_contract_backed_markdown_to_word_unsafe_filename(self):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"))
            response = executor.run(
                "agency.document.markdown-to-word",
                {
                    "markdown_text": "# Report",
                    "filename": "../report.docx",
                    "img_directory": "reports",
                },
                actor="approved/user-runtime",
            )

            self.assertEqual(response.verdict, "deny")
            self.assertIsNone(response.result)
            self.assertIn("filename must be a safe document name", response.errors[0])

    def test_runtime_runs_contract_backed_excel_text_writer(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = _create_workbook(root / "results.xlsx")
            text_path = root / "result.txt"
            text_path.write_text("contract spreadsheet", encoding="utf-8")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root)]),
                run_store=JsonlToolRunStore(root / "tool_runs.jsonl"),
            )

            response = executor.run(
                "agency.excel.write-text",
                {
                    "sheet_name": "Sheet1",
                    "excel_file_path": str(workbook_path),
                    "text_file_path": str(text_path),
                    "serial_number": 1,
                    "header_title": "Notes",
                },
                actor="approved/user-runtime",
            )

            worksheet = load_workbook(workbook_path).active
            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status"], "success")
            self.assertEqual(worksheet["A1"].value, "Notes")
            self.assertEqual(worksheet["A2"].value, "contract spreadsheet")
            self.assertEqual(response.filesChanged[0].path, str(workbook_path))

    def test_runtime_runs_contract_backed_excel_json_writer(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = _create_workbook(root / "results.xlsx")
            json_path = root / "result.json"
            json_path.write_text(json.dumps({"Summary": "OK"}), encoding="utf-8")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root)]),
                run_store=JsonlToolRunStore(root / "tool_runs.jsonl"),
            )

            response = executor.run(
                "agency.excel.write-json",
                {
                    "sheet_name": "Sheet1",
                    "excel_file_path": str(workbook_path),
                    "json_file_path": str(json_path),
                    "serial_number": 1,
                },
                actor="approved/user-runtime",
            )

            worksheet = load_workbook(workbook_path).active
            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status"], "success")
            self.assertEqual(worksheet["A1"].value, "Summary")
            self.assertEqual(worksheet["A2"].value, "OK")

    def test_runtime_runs_contract_backed_excel_image_writer(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = _create_workbook(root / "results.xlsx")
            image_path = root / "sample.png"
            Image.new("RGB", (20, 20), color="red").save(image_path)
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root)]),
                run_store=JsonlToolRunStore(root / "tool_runs.jsonl"),
            )

            response = executor.run(
                "agency.excel.write-image",
                {
                    "sheet_name": "Sheet1",
                    "excel_file_path": str(workbook_path),
                    "image_path": str(image_path),
                    "serial_number": 1,
                    "header_title": "Evidence",
                },
                actor="approved/user-runtime",
            )

            worksheet = load_workbook(workbook_path).active
            self.assertEqual(response.verdict, "ok")
            self.assertEqual(response.result["status"], "success")
            self.assertEqual(worksheet["A1"].value, "Evidence")

    def test_runtime_denies_contract_backed_excel_writer_outside_allowlist(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = _create_workbook(root / "blocked" / "results.xlsx")
            text_path = root / "result.txt"
            text_path.write_text("blocked", encoding="utf-8")
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_file_write_dirs=[str(root / "allowed")]),
                run_store=JsonlToolRunStore(root / "tool_runs.jsonl"),
            )

            response = executor.run(
                "agency.excel.write-text",
                {
                    "sheet_name": "Sheet1",
                    "excel_file_path": str(workbook_path),
                    "text_file_path": str(text_path),
                    "serial_number": 1,
                },
                actor="approved/user-runtime",
            )

            self.assertEqual(response.verdict, "deny")
            self.assertIsNone(response.result)
            self.assertIn("not under an allowlisted directory", response.errors[0])

    def test_runtime_denies_before_patch_validation_when_policy_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            executor = ToolRuntimeExecutor(
                policy_engine=PolicyEngine(allowed_repos=[]),
                run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
            )
            response = executor.run(
                "sandbox-edit",
                {
                    "repo": "/tmp/not-allowed",
                    "ref": "main",
                    "changes": [{"path": "README.md", "patch": "not a patch"}],
                    "dryRun": True,
                },
            )

        self.assertEqual(response.verdict, "deny")
        self.assertIsNone(response.patch)
        self.assertTrue(response.errors)

    def test_api_exposes_contracts_and_runtime_validation(self):
        client = TestClient(create_app(context=create_test_api_context()))

        list_response = client.get("/tools/contracts")
        self.assertEqual(list_response.status_code, 200)
        self.assertIn("sandbox-edit", {item["name"] for item in list_response.json()["items"]})
        self.assertIn("agency.tool.list", {item["name"] for item in list_response.json()["items"]})

        get_response = client.get("/tools/contracts/sandbox-edit")
        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(get_response.json()["name"], "sandbox-edit")
        self.assertIn("riskLabels", get_response.json())
        self.assertIn("filesystem", get_response.json()["riskLabels"])

        invalid_run = client.post("/tools/sandbox-edit/run", json={"repo": "/tmp/example"})
        self.assertEqual(invalid_run.status_code, 400)
        self.assertIn("validation failed", invalid_run.json()["detail"])

        capabilities = client.get("/capabilities")
        self.assertEqual(capabilities.status_code, 200)
        body = capabilities.json()
        self.assertEqual(body["name"], "agency-runtime")
        self.assertEqual(
            [tool.id for tool in list_builtin_tool_definitions() if tool.id not in {item["name"] for item in body["tools"]}],
            [],
        )
        self.assertIn("sandbox-edit", {item["name"] for item in body["tools"]})
        self.assertIn("agency.http.request", {item["name"] for item in body["tools"]})
        self.assertIn("agency.workflow.list", {item["name"] for item in body["tools"]})
        self.assertIn("agency.workflow.get", {item["name"] for item in body["tools"]})
        self.assertIn("agency.workflow.propose-create", {item["name"] for item in body["tools"]})
        self.assertIn("agency.workflow.propose-update", {item["name"] for item in body["tools"]})
        self.assertIn("agency.tool.get", {item["name"] for item in body["tools"]})
        self.assertIn("agency.tool.propose-create", {item["name"] for item in body["tools"]})
        self.assertIn("agency.tool.propose-update", {item["name"] for item in body["tools"]})
        self.assertIn("agency.memory.list", {item["name"] for item in body["tools"]})
        self.assertIn("agency.memory.remember", {item["name"] for item in body["tools"]})
        self.assertIn("agency.memory.update", {item["name"] for item in body["tools"]})
        self.assertIn("agency.memory.delete", {item["name"] for item in body["tools"]})
        self.assertIn("agency.tool.list", {item["name"] for item in body["tools"]})
        self.assertIn("agency.command.run", {item["name"] for item in body["tools"]})
        self.assertIn("agency.file.write-text", {item["name"] for item in body["tools"]})
        self.assertIn("agency.document.markdown-to-word", {item["name"] for item in body["tools"]})
        self.assertIn("agency.excel.write-text", {item["name"] for item in body["tools"]})
        self.assertIn("agency.excel.write-json", {item["name"] for item in body["tools"]})
        self.assertIn("agency.excel.write-image", {item["name"] for item in body["tools"]})
        modules = body["modules"]
        self.assertNotIn("smart_home", modules)
        self.assertNotIn("physical_devices", modules)
        self.assertNotIn("home_assistant.call_service", {item["name"] for item in body["tools"]})
        self.assertNotIn("agency.device.list", {item["name"] for item in body["tools"]})
        self.assertNotIn("agency.device.command", {item["name"] for item in body["tools"]})
        capability_by_name = {item["name"]: item for item in body["tools"]}
        workflow_run_execution = capability_by_name["agency.workflow.run"]["execution"]
        self.assertEqual(workflow_run_execution["executionMode"], "approval_context")
        self.assertTrue(workflow_run_execution["supportsApprovalRequest"])
        self.assertIn("conversation_id", workflow_run_execution["inputContextFields"])
        self.assertIn("workflow_execution", workflow_run_execution["sideEffects"])
        proposal_execution = capability_by_name["agency.workflow.propose-create"]["execution"]
        self.assertEqual(proposal_execution["executionMode"], "conversation_context")
        self.assertTrue(proposal_execution["requiresConversation"])
        self.assertIn("approval_request", proposal_execution["sideEffects"])
        browser_execution = capability_by_name["agency.browser.click"]["execution"]
        self.assertEqual(browser_execution["executionMode"], "direct")
        self.assertIn("browser_mutation", browser_execution["sideEffects"])
        self.assertTrue(browser_execution["policyNotes"])
        memory_execution = capability_by_name["agency.memory.remember"]["execution"]
        self.assertEqual(memory_execution["executionMode"], "api_context")
        self.assertIn("memory", memory_execution["sideEffects"])
        self.assertEqual(body["events"]["streamUrl"], "/api/runtime/events/stream")

    def test_capabilities_exposes_optional_module_availability_contract(self):
        try:
            with patch.dict(
                "os.environ",
                {
                    "AGENCY_OPTIONAL_MODULE_SPEC_REFS": "tests.fixtures.optional_module_pack:external_home_pack_specs",
                    "SMART_HOME_MODULE_ENABLED": "false",
                    "PHYSICAL_DEVICES_MODULE_ENABLED": "false",
                },
                clear=False,
            ):
                reset_settings_cache()
                get_default_contract_registry.cache_clear()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.get("/capabilities")

            self.assertEqual(response.status_code, 200)
            modules = response.json()["modules"]
            self.assertFalse(modules["smart_home"]["available"])
            self.assertEqual(modules["smart_home"]["status"], "disabled")
            self.assertEqual(modules["smart_home"]["routePrefix"], "/api/smart-home")
            self.assertEqual(modules["smart_home"]["tools"]["vendorSpecific"], [])
            self.assertIn(
                "home_assistant.call_service",
                modules["smart_home"]["hiddenWhenUnavailable"]["toolNames"],
            )
            self.assertIn("/api/smart-home", modules["smart_home"]["hiddenWhenUnavailable"]["routePrefixes"])
            self.assertFalse(modules["physical_devices"]["available"])
            self.assertEqual(modules["physical_devices"]["status"], "disabled")
            self.assertEqual(modules["physical_devices"]["routePrefix"], "/api/devices")
            self.assertEqual(modules["physical_devices"]["eventRoutePrefix"], "/api/physical/events")
            self.assertEqual(modules["physical_devices"]["tools"]["preferred"], [])
            self.assertIn(
                "agency.device.command",
                modules["physical_devices"]["hiddenWhenUnavailable"]["toolNames"],
            )
            tool_names = {item["name"] for item in response.json()["tools"]}
            self.assertNotIn("agency.device.list", tool_names)
            self.assertNotIn("agency.physical.event-bus.health", tool_names)

            with patch.dict(
                "os.environ",
                {
                    "AGENCY_OPTIONAL_MODULE_SPEC_REFS": "tests.fixtures.optional_module_pack:external_home_pack_specs",
                    "SMART_HOME_MODULE_ENABLED": "true",
                    "PHYSICAL_DEVICES_MODULE_ENABLED": "true",
                },
                clear=False,
            ):
                reset_settings_cache()
                get_default_contract_registry.cache_clear()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.get("/capabilities")

            self.assertEqual(response.status_code, 200)
            modules = response.json()["modules"]
            self.assertTrue(modules["smart_home"]["available"])
            self.assertEqual(modules["smart_home"]["status"], "available")
            self.assertIn("home_assistant.call_service", modules["smart_home"]["tools"]["vendorSpecific"])
            self.assertTrue(modules["physical_devices"]["available"])
            self.assertEqual(modules["physical_devices"]["status"], "available")
            self.assertIn("agency.device.list", modules["physical_devices"]["tools"]["readOnly"])
            self.assertIn("agency.device.command", modules["physical_devices"]["tools"]["mutating"])
        finally:
            reset_settings_cache()
            get_default_contract_registry.cache_clear()

    def test_api_runs_contract_backed_tool_list(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post("/tools/agency.tool.list/run", json={})

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertEqual(body["policyVerdict"]["score"], 0)
                self.assertEqual(body["result"]["count"], len(list_builtin_tool_definitions()))
                self.assertIn("agency.speech.listen", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.speech.speak", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.speech.continue", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.context", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.search", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.expand", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.neighbors", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.path", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.graph.summarize-subgraph", {item["id"] for item in body["result"]["items"]})
                self.assertIn("agency.tool.list", {item["id"] for item in body["result"]["items"]})
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    @patch("app.tools.runtime.executor.open_browser")
    def test_api_runs_contract_backed_browser_open(self, mock_open_browser):
        mock_open_browser.return_value = {
            "url": "https://example.test",
            "title": "Example",
            "runtime_root": "/tmp/browser-runtime",
            "message": "Browser started.",
        }
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post("/tools/agency.browser.open/run", json={"url": "https://example.test"})

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertEqual(body["result"]["status"], "ok")
                self.assertEqual(body["result"]["output"]["title"], "Example")
                self.assertTrue(body["signature"].startswith("sha256:"))
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_runs_contract_backed_workflow_run(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = create_test_api_context()
                asyncio.run(
                    context.user_repo.create(
                        UserDefinition(
                            id="user-contract-run",
                            email="contract-run@example.com",
                            display_name="Contract Runner",
                        )
                    )
                )
                workflow = _workflow_definition("workflow-contract-run").model_copy(
                    update={"metadata": {"created_by": "user-contract-run", "owner_ids": ["user-contract-run"]}}
                )
                asyncio.run(context.workflow_repo.create(workflow))
                client = TestClient(create_app(context=context))
                client.headers.update(
                    {"x-agency-user-id": "user-contract-run", "x-agency-user-email": "contract-run@example.com"}
                )
                response = client.post(
                    "/tools/agency.workflow.run/run",
                    json={"workflow_id": "workflow-contract-run", "input_payload": {"topic": "launch"}},
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertEqual(body["result"]["workflow_id"], "workflow-contract-run")
                self.assertEqual(body["result"]["execution_status"], "queued")
                self.assertTrue(body["result"]["execution_id"])
                self.assertTrue(body["signature"].startswith("sha256:"))
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_marks_protected_workflow_run_as_approval_context_required(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = create_test_api_context()
                asyncio.run(
                    context.user_repo.create(
                        UserDefinition(
                            id="user-protected-run",
                            email="protected-run@example.com",
                            display_name="Protected Runner",
                        )
                    )
                )
                workflow = _workflow_definition("workflow-protected-contract-run").model_copy(
                    update={
                        "metadata": {
                            "protected_execution": True,
                            "created_by": "user-protected-run",
                            "owner_ids": ["user-protected-run"],
                        }
                    }
                )
                asyncio.run(context.workflow_repo.create(workflow))
                client = TestClient(create_app(context=context))
                client.headers.update(
                    {
                        "x-agency-user-id": "user-protected-run",
                        "x-agency-user-email": "protected-run@example.com",
                    }
                )
                response = client.post(
                    "/tools/agency.workflow.run/run",
                    json={"workflow_id": "workflow-protected-contract-run"},
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "warn")
                self.assertEqual(body["result"]["status"], "requires_approval_context")
                self.assertTrue(body["signature"].startswith("sha256:"))
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_requests_protected_workflow_approval_with_conversation_context(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = asyncio.run(_prepare_conversation_context("conversation-contract-run"))
                asyncio.run(
                    context.user_repo.create(
                        UserDefinition(
                            id="user-contract",
                            email="contract@example.com",
                            display_name="Contract User",
                        )
                    )
                )
                workflow = _workflow_definition("workflow-protected-approval-contract").model_copy(
                    update={
                        "metadata": {
                            "protected_execution": True,
                            "created_by": "user-contract",
                            "owner_ids": ["user-contract"],
                        }
                    }
                )
                asyncio.run(context.workflow_repo.create(workflow))
                client = TestClient(create_app(context=context))
                client.headers.update(
                    {"x-agency-user-id": "user-contract", "x-agency-user-email": "contract@example.com"}
                )
                response = client.post(
                    "/tools/agency.workflow.run/run",
                    json={
                        "workflow_id": "workflow-protected-approval-contract",
                        "input_payload": {"topic": "approval"},
                        "conversation_id": "conversation-contract-run",
                    },
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertEqual(body["result"]["status"], "approval_requested")
                self.assertEqual(body["result"]["approval_request"]["approval_type"], "workflow_execution")
                self.assertEqual(body["result"]["approval_request"]["target_id"], "workflow-protected-approval-contract")
                self.assertTrue(body["signature"].startswith("sha256:"))
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_creates_workflow_proposal_approval_with_conversation_context(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = asyncio.run(_prepare_conversation_context("conversation-contract-proposal"))
                client = TestClient(create_app(context=context))
                response = client.post(
                    "/tools/agency.workflow.propose-create/run",
                    json={
                        "summary": "Create contract proposal workflow.",
                        "workflow": _workflow_payload("workflow-contract-proposal"),
                        "conversation_id": "conversation-contract-proposal",
                    },
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertEqual(body["result"]["status"], "approval_requested")
                self.assertEqual(body["result"]["approval_request"]["approval_type"], "workflow_create")
                self.assertEqual(body["result"]["approval_request"]["target_id"], "workflow-contract-proposal")
                self.assertTrue(body["signature"].startswith("sha256:"))
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_api_runs_contract_backed_http_request(self, mock_execute):
        mock_execute.return_value = {"status_code": 201, "response": {"created": True}}
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        previous_allowed_hosts = os.environ.get("TOOL_HTTP_ALLOWED_HOSTS")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                os.environ["TOOL_HTTP_ALLOWED_HOSTS"] = "api.example.test"
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post(
                    "/tools/agency.http.request/run",
                    json={"url": "https://api.example.test/items", "method": "POST", "body": {"name": "item"}},
                )

                self.assertEqual(response.status_code, 401)
                mock_execute.assert_not_called()
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            if previous_allowed_hosts is None:
                os.environ.pop("TOOL_HTTP_ALLOWED_HOSTS", None)
            else:
                os.environ["TOOL_HTTP_ALLOWED_HOSTS"] = previous_allowed_hosts
            reset_settings_cache()

    def test_api_runs_context_backed_workflow_list_get_and_tool_get(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = create_test_api_context()
                asyncio.run(
                    context.user_repo.create(
                        UserDefinition(
                            id="user-execution-contract",
                            email="execution-contract@example.com",
                            display_name="Execution Contract User",
                        )
                    )
                )
                workflow = asyncio.run(context.workflow_repo.create(_workflow_definition("workflow-api")))
                execution = Execution(
                    id="execution-api-eval",
                    workflow_id=workflow.id,
                    runtime_adapter_id="native",
                    status=ExecutionStatus.COMPLETED,
                    input_payload={"topic": "contracts"},
                    output_payload={"final_output": "done"},
                    created_by="user-execution-contract",
                )
                asyncio.run(context.execution_store.save_execution(execution))
                asyncio.run(
                    context.execution_store.save_event(
                        ExecutionEvent(
                            execution_id=execution.id,
                            workflow_id=workflow.id,
                            event_type=ExecutionEventType.EXECUTION_COMPLETED,
                            sequence=1,
                            payload={"output": execution.output_payload},
                        )
                    )
                )
                asyncio.run(
                    context.execution_store.save_artifact(
                        ExecutionArtifact(
                            id="artifact-api-eval",
                            execution_id=execution.id,
                            artifact_type="text",
                            name="result.txt",
                            content_text="artifact body",
                        )
                    )
                )
                asyncio.run(context.ensure_builtin_tool_seed_data())
                client = TestClient(create_app(context=context))
                client.headers.update(
                    {
                        "x-agency-user-id": "user-execution-contract",
                        "x-agency-user-email": "execution-contract@example.com",
                    }
                )

                list_response = client.post("/tools/agency.workflow.list/run", json={})
                get_response = client.post("/tools/agency.workflow.get/run", json={"workflow_id": workflow.id})
                tool_response = client.post("/tools/agency.tool.get/run", json={"tool_id": "agency.http.request"})
                execution_list_response = client.post(
                    "/tools/agency.execution.list/run",
                    json={"workflow_id": workflow.id, "status": ["completed"]},
                )
                execution_response = client.post(
                    "/tools/agency.execution.get/run",
                    json={"execution_id": execution.id},
                )
                events_response = client.post(
                    "/tools/agency.execution.events/run",
                    json={"execution_id": execution.id, "event_types": ["execution.completed"]},
                )
                artifacts_response = client.post(
                    "/tools/agency.execution.artifacts/run",
                    json={"execution_id": execution.id, "include_content": True},
                )

                self.assertEqual(list_response.status_code, 200)
                self.assertEqual(list_response.json()["result"]["workflows"][0]["id"], workflow.id)
                self.assertEqual(get_response.status_code, 200)
                self.assertEqual(get_response.json()["result"]["workflow"]["id"], workflow.id)
                self.assertEqual(tool_response.status_code, 200)
                self.assertEqual(tool_response.json()["result"]["tool"]["id"], "agency.http.request")
                self.assertEqual(execution_list_response.status_code, 200)
                self.assertEqual(execution_list_response.json()["result"]["items"][0]["id"], execution.id)
                self.assertEqual(execution_response.status_code, 200)
                self.assertEqual(execution_response.json()["result"]["execution"]["id"], execution.id)
                self.assertEqual(events_response.status_code, 200)
                self.assertEqual(events_response.json()["result"]["count"], 1)
                self.assertEqual(artifacts_response.status_code, 200)
                self.assertEqual(artifacts_response.json()["result"]["items"][0]["content_text"], "artifact body")
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_execution_contract_runtime_enforces_owner_and_admin_access(self):
        async def run_assertions() -> None:
            context = create_test_api_context()
            for user in (
                UserDefinition(
                    id="user-execution-owner",
                    email="execution-owner@example.com",
                    display_name="Execution Owner",
                ),
                UserDefinition(
                    id="user-execution-other",
                    email="execution-other@example.com",
                    display_name="Other Execution User",
                ),
                UserDefinition(
                    id="user-execution-admin",
                    email="execution-admin@example.com",
                    display_name="Execution Admin",
                    roles=["admin"],
                ),
            ):
                await context.user_repo.create(user)

            workflow = await context.workflow_repo.create(_workflow_definition("workflow-execution-authz"))
            foreign_workflow = await context.workflow_repo.create(
                _workflow_definition("workflow-foreign-authz").model_copy(
                    update={
                        "metadata": {
                            "created_by": "user-execution-other",
                            "owner_ids": ["user-execution-other"],
                        }
                    }
                )
            )
            owner_execution = Execution(
                id="execution-owner-visible",
                workflow_id=workflow.id,
                runtime_adapter_id="native",
                status=ExecutionStatus.COMPLETED,
                input_payload={},
                created_by="user-execution-owner",
            )
            foreign_execution = Execution(
                id="execution-foreign-hidden",
                workflow_id=workflow.id,
                runtime_adapter_id="native",
                status=ExecutionStatus.PAUSED,
                input_payload={"secret": "foreign-input"},
                created_by="user-execution-other",
            )
            await context.execution_store.save_execution(owner_execution)
            await context.execution_store.save_execution(foreign_execution)
            await context.execution_store.save_event(
                ExecutionEvent(
                    execution_id=foreign_execution.id,
                    workflow_id=workflow.id,
                    event_type=ExecutionEventType.EXECUTION_STARTED,
                    sequence=1,
                    payload={"secret": "foreign-event"},
                )
            )
            await context.execution_store.save_artifact(
                ExecutionArtifact(
                    id="artifact-foreign-hidden",
                    execution_id=foreign_execution.id,
                    artifact_type="text",
                    name="foreign.txt",
                    content_text="foreign artifact",
                )
            )

            context.control_plane.pause = AsyncMock(return_value=foreign_execution)
            context.control_plane.resume = AsyncMock(return_value=foreign_execution)
            context.control_plane.cancel = AsyncMock(return_value=foreign_execution)

            with tempfile.TemporaryDirectory() as tmp:
                executor = ToolRuntimeExecutor(
                    context=context,
                    run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                )
                owner_list = await executor.run_async(
                    "agency.execution.list", {}, actor="user-execution-owner"
                )
                self.assertEqual(
                    [item["id"] for item in owner_list.result["items"]],
                    [owner_execution.id],
                )
                foreign_workflow_run = await executor.run_async(
                    "agency.workflow.run",
                    {"workflow_id": foreign_workflow.id},
                    actor="user-execution-owner",
                )
                self.assertEqual(foreign_workflow_run.result["status"], "error")
                self.assertIn("not found", foreign_workflow_run.result["error"])
                self.assertEqual(len(await context.execution_store.list_executions()), 2)

                foreign_read_cases = (
                    ("agency.execution.get", {"execution_id": foreign_execution.id}),
                    ("agency.execution.events", {"execution_id": foreign_execution.id}),
                    ("agency.execution.artifacts", {"execution_id": foreign_execution.id}),
                )
                for tool_name, payload in foreign_read_cases:
                    response = await executor.run_async(tool_name, payload, actor="user-execution-owner")
                    self.assertEqual(response.result["status"], "error", tool_name)
                    self.assertIn("not found", response.result["error"], tool_name)

                foreign_control_cases = (
                    ("agency.execution.pause", {"execution_id": foreign_execution.id}),
                    ("agency.execution.resume", {"execution_id": foreign_execution.id}),
                    ("agency.execution.cancel", {"execution_id": foreign_execution.id}),
                )
                for tool_name, payload in foreign_control_cases:
                    response = await executor.run_async(tool_name, payload, actor="user-execution-owner")
                    self.assertEqual(response.result["status"], "error", tool_name)
                context.control_plane.pause.assert_not_awaited()
                context.control_plane.resume.assert_not_awaited()
                context.control_plane.cancel.assert_not_awaited()

                with (
                    patch.object(
                        ExecutionService,
                        "list_execution_approvals",
                        new_callable=AsyncMock,
                    ) as list_approvals,
                    patch.object(ExecutionService, "approve", new_callable=AsyncMock) as approve,
                    patch.object(ExecutionService, "reject", new_callable=AsyncMock) as reject,
                ):
                    approval_cases = (
                        ("agency.execution.approvals", {"execution_id": foreign_execution.id}),
                        (
                            "agency.execution.approve",
                            {"execution_id": foreign_execution.id, "tool_id": "agency.command.run"},
                        ),
                        (
                            "agency.execution.reject",
                            {"execution_id": foreign_execution.id, "tool_id": "agency.command.run"},
                        ),
                    )
                    for tool_name, payload in approval_cases:
                        response = await executor.run_async(tool_name, payload, actor="user-execution-owner")
                        self.assertEqual(response.result["status"], "error", tool_name)
                    list_approvals.assert_not_awaited()
                    approve.assert_not_awaited()
                    reject.assert_not_awaited()

                admin_list = await executor.run_async(
                    "agency.execution.list", {}, actor="user-execution-admin"
                )
                self.assertEqual(
                    {item["id"] for item in admin_list.result["items"]},
                    {owner_execution.id, foreign_execution.id},
                )
                for tool_name in (
                    "agency.execution.get",
                    "agency.execution.events",
                    "agency.execution.artifacts",
                ):
                    admin_response = await executor.run_async(
                        tool_name,
                        {"execution_id": foreign_execution.id},
                        actor="user-execution-admin",
                    )
                    self.assertEqual(admin_response.result["status"], "ok", tool_name)

                admin_pause = await executor.run_async(
                    "agency.execution.pause",
                    {"execution_id": foreign_execution.id},
                    actor="user-execution-admin",
                )
                self.assertEqual(admin_pause.result["status"], "ok")
                context.control_plane.pause.assert_awaited_once_with(foreign_execution.id)

        asyncio.run(run_assertions())

    def test_execution_contract_http_route_requires_authenticated_actor(self):
        context = create_test_api_context()
        client = TestClient(create_app(context=context))

        response = client.post("/tools/agency.execution.list/run", json={})

        self.assertEqual(response.status_code, 401)

    def test_api_marks_direct_proposal_tool_as_conversation_context_required(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post(
                    "/tools/agency.workflow.propose-create/run",
                    json={"goal": "Create a workflow that drafts a report."},
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "warn")
                self.assertEqual(body["result"]["status"], "requires_conversation_context")
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_runs_context_backed_memory_crud(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        headers = {"x-agency-user-id": "user-memory", "x-agency-user-email": "memory@example.com"}
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = create_test_api_context()
                client = TestClient(create_app(context=context))
                client.post("/users/sync", json={"id": "user-memory", "email": "memory@example.com"})

                remember = client.post(
                    "/tools/agency.memory.remember/run",
                    headers=headers,
                    json={"scope": "user", "content": "Remember API contract memory.", "confirmed": True},
                )
                memory_id = remember.json()["result"]["memory"]["id"]
                listed = client.post(
                    "/tools/agency.memory.list/run",
                    headers=headers,
                    json={"scope": "user", "query": "contract"},
                )
                updated = client.post(
                    "/tools/agency.memory.update/run",
                    headers=headers,
                    json={"memory_id": memory_id, "summary": "API contract memory."},
                )
                deleted = client.post(
                    "/tools/agency.memory.delete/run",
                    headers=headers,
                    json={"memory_id": memory_id},
                )

                self.assertEqual(remember.status_code, 200)
                self.assertEqual(remember.json()["result"]["memory"]["created_by_user_id"], "user-memory")
                self.assertEqual(listed.json()["result"]["memories"][0]["id"], memory_id)
                self.assertEqual(updated.json()["result"]["memory"]["summary"], "API contract memory.")
                self.assertTrue(deleted.json()["result"]["deleted"])
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_rejects_header_asserted_command_approval(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                context = create_test_api_context()
                client = TestClient(create_app(context=context))
                client.post(
                    "/users/sync",
                    json={
                        "id": "user-command",
                        "email": "command@example.com",
                        "display_name": "Command User",
                    },
                )
                marker = Path(tmp) / "header-approval-bypass"
                for asserted_approval in ("true", "1", "yes", "approved", "TRUE"):
                    with self.subTest(asserted_approval=asserted_approval):
                        response = client.post(
                            "/tools/agency.command.run/run",
                            headers={
                                "x-agency-user-id": "user-command",
                                "x-agency-user-email": "command@example.com",
                                "x-agency-command-approved": asserted_approval,
                            },
                            json={
                                "command": f"printf bypassed > {marker}",
                                "mode": "bash",
                                "timeout_seconds": 2,
                            },
                        )

                        self.assertEqual(response.status_code, 403)
                        self.assertIn("execution-bound approval", response.json()["detail"])
                        self.assertFalse(marker.exists())

                anonymous = client.post(
                    "/tools/agency.command.run/run",
                    headers={"x-agency-command-approved": "true"},
                    json={"command": f"printf bypassed > {marker}", "mode": "bash"},
                )
                self.assertEqual(anonymous.status_code, 401)
                self.assertFalse(marker.exists())
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_blocks_direct_file_write_text_before_mutation(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        previous_allowed_dirs = os.environ.get("TOOL_FILE_WRITE_ALLOWED_DIRS")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                root = Path(tmp)
                os.environ["TOOL_RUN_STORE_PATH"] = str(root / "api_tool_runs.jsonl")
                os.environ["TOOL_FILE_WRITE_ALLOWED_DIRS"] = str(root)
                reset_settings_cache()
                context = create_test_api_context()
                asyncio.run(
                    context.user_repo.create(
                        UserDefinition(
                            id="user-file-write",
                            email="file-write@example.com",
                            display_name="File Write User",
                        )
                    )
                )
                client = TestClient(create_app(context=context))
                payload = {
                    "base_folder": str(root),
                    "filename": "api.txt",
                    "content": "api file",
                    "mode": "write",
                }
                anonymous_response = client.post(
                    "/tools/agency.file.write-text/run",
                    json=payload,
                )
                authenticated_response = client.post(
                    "/tools/agency.file.write-text/run",
                    headers={
                        "x-agency-user-id": "user-file-write",
                        "x-agency-user-email": "file-write@example.com",
                    },
                    json=payload,
                )

                self.assertEqual(anonymous_response.status_code, 401)
                self.assertEqual(authenticated_response.status_code, 403)
                self.assertIn("execution-bound approval", authenticated_response.json()["detail"])
                self.assertFalse((root / "api.txt").exists())
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            if previous_allowed_dirs is None:
                os.environ.pop("TOOL_FILE_WRITE_ALLOWED_DIRS", None)
            else:
                os.environ["TOOL_FILE_WRITE_ALLOWED_DIRS"] = previous_allowed_dirs
            reset_settings_cache()

    @patch("app.tools.implementations.documents.upload_to_s3")
    def test_api_runs_contract_backed_markdown_to_word(self, mock_upload):
        mock_upload.return_value = {"uploaded_files": ["user_api/workflow_reports/run_proc-api/report.docx"]}
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                os.environ["TOOL_RUN_STORE_PATH"] = str(Path(tmp) / "api_tool_runs.jsonl")
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post(
                    "/tools/agency.document.markdown-to-word/run",
                    json={
                        "markdown_text": "# API Report\n\nBody",
                        "filename": "report.docx",
                        "img_directory": "reports",
                        "process_id": "proc-api",
                        "run_by": "api",
                    },
                )

                self.assertEqual(response.status_code, 401)
                mock_upload.assert_not_called()
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_api_runs_contract_backed_excel_text_writer(self):
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        previous_allowed_dirs = os.environ.get("TOOL_FILE_WRITE_ALLOWED_DIRS")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                root = Path(tmp)
                workbook_path = _create_workbook(root / "results.xlsx")
                text_path = root / "result.txt"
                text_path.write_text("api spreadsheet", encoding="utf-8")
                os.environ["TOOL_RUN_STORE_PATH"] = str(root / "api_tool_runs.jsonl")
                os.environ["TOOL_FILE_WRITE_ALLOWED_DIRS"] = str(root)
                reset_settings_cache()
                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post(
                    "/tools/agency.excel.write-text/run",
                    json={
                        "sheet_name": "Sheet1",
                        "excel_file_path": str(workbook_path),
                        "text_file_path": str(text_path),
                        "serial_number": 1,
                        "header_title": "Notes",
                    },
                )

                self.assertEqual(response.status_code, 401)
                self.assertIsNone(load_workbook(workbook_path).active["A2"].value)
        finally:
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            if previous_allowed_dirs is None:
                os.environ.pop("TOOL_FILE_WRITE_ALLOWED_DIRS", None)
            else:
                os.environ["TOOL_FILE_WRITE_ALLOWED_DIRS"] = previous_allowed_dirs
            reset_settings_cache()

    def test_api_runs_sandbox_edit_dry_run_against_allowlisted_repo(self):
        previous_allowlist = os.environ.get("SANDBOX_EDIT_ALLOWED_REPOS")
        previous_store = os.environ.get("TOOL_RUN_STORE_PATH")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                repo = Path(tmp)
                subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True, text=True)
                (repo / "README.md").write_text("hello\n", encoding="utf-8")
                os.environ["SANDBOX_EDIT_ALLOWED_REPOS"] = str(repo)
                os.environ["TOOL_RUN_STORE_PATH"] = str(repo / "api_tool_runs.jsonl")
                reset_settings_cache()

                client = TestClient(create_app(context=create_test_api_context()))
                response = client.post(
                    "/tools/sandbox-edit/run",
                    json={
                        "repo": str(repo),
                        "ref": "main",
                        "changes": [{"path": "README.md", "patch": README_PATCH}],
                        "dryRun": True,
                    },
                )

                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertEqual(body["verdict"], "ok")
                self.assertTrue(body["signature"].startswith("sha256:"))
                self.assertEqual(body["filesChanged"][0]["path"], "README.md")
                self.assertEqual((repo / "README.md").read_text(encoding="utf-8"), "hello\n")
                records = JsonlToolRunStore(repo / "api_tool_runs.jsonl").list_records()
                self.assertEqual(len(records), 1)
                self.assertEqual(records[0].tool_name, "sandbox-edit")
        finally:
            if previous_allowlist is None:
                os.environ.pop("SANDBOX_EDIT_ALLOWED_REPOS", None)
            else:
                os.environ["SANDBOX_EDIT_ALLOWED_REPOS"] = previous_allowlist
            if previous_store is None:
                os.environ.pop("TOOL_RUN_STORE_PATH", None)
            else:
                os.environ["TOOL_RUN_STORE_PATH"] = previous_store
            reset_settings_cache()

    def test_build_dry_run_pr_payload_is_external_integration_ready(self):
        payload = build_dry_run_pr_payload(
            repo="git@example.com:agency/agency-fe.git",
            branch="agent/1234-sandbox",
            base="main",
            title="[dry-run] propose change",
            files=[{"path": "README.md", "patch": README_PATCH}],
            metadata={"source": "test"},
            agent="codex/test",
        )

        self.assertEqual(payload["action"], "create_pr_dry_run")
        self.assertEqual(payload["patchJson"]["files"][0]["path"], "README.md")
        self.assertEqual(payload["metadata"]["agent"], "codex/test")
        self.assertEqual(payload["metadata"]["source"], "test")

    def test_runtime_publishes_tool_lifecycle_events(self):
        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with tempfile.TemporaryDirectory() as tmp:
                    repo = Path(tmp)
                    subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True, text=True)
                    (repo / "README.md").write_text("hello\n", encoding="utf-8")
                    executor = ToolRuntimeExecutor(
                        policy_engine=PolicyEngine(allowed_repos=[str(repo)]),
                        run_store=JsonlToolRunStore(repo / "tool_runs.jsonl"),
                    )

                    response = executor.run(
                        "sandbox-edit",
                        {
                            "repo": str(repo),
                            "ref": "main",
                            "changes": [{"path": "README.md", "patch": README_PATCH}],
                            "dryRun": True,
                        },
                        actor="codex/test",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(response.verdict, "ok")
                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    semantic_types = [event.metadata.get("semanticType") for event in events]
                    self.assertIn("tool.run.started", semantic_types)
                    self.assertIn("tool.policy.completed", semantic_types)
                    self.assertIn("tool.run.completed", semantic_types)
                    self.assertIn("ok", [event.metadata.get("verdict") for event in events])
                    started_event = events[semantic_types.index("tool.run.started")]
                    self.assertIn("filesystem", started_event.metadata["riskLabels"])
                    self.assertEqual(started_event.metadata["localPrivilegedExecution"], True)
            finally:
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_publishes_onecli_http_observability_events(self, mock_execute):
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with tempfile.TemporaryDirectory() as tmp, patch.dict(
                        os.environ,
                        {
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                            "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_AGENT_TOKEN",
                            "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                            "ONECLI_AGENT_TOKEN": "test-onecli-agent-token",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    executor = ToolRuntimeExecutor(
                        policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                        run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                    )
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                        },
                        actor="user-runtime",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(response.verdict, "ok")
                    correlation_id = response.result["onecli"]["correlation_id"]
                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    semantic_types = [event.metadata.get("semanticType") for event in onecli_events]
                    self.assertEqual(
                        semantic_types,
                        ["onecli.http.request.started", "onecli.http.request.completed"],
                    )
                    for event in onecli_events:
                        self.assertEqual(event.metadata["correlation_id"], correlation_id)
                        self.assertEqual(event.metadata["target_host"], "api.example.test")
                        self.assertEqual(event.metadata["gateway_mode"], "proxy")
                        self.assertEqual(event.metadata["agent_identity"]["agency_actor"], "user-runtime")
                        self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(event.metadata))
                    self.assertEqual(onecli_events[-1].metadata["status_code"], 200)
                    kwargs = mock_execute.call_args.kwargs
                    self.assertEqual(
                        kwargs["headers"]["X-Agency-OneCLI-Correlation-ID"],
                        correlation_id,
                    )
                    self.assertEqual(kwargs["headers"]["X-Agency-User-ID"], "user-runtime")
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_runtime_publishes_onecli_http_rate_limited_event(self, mock_execute):
        mock_execute.return_value = {"status_code": 429, "response": {"error": "rate limited"}}

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with tempfile.TemporaryDirectory() as tmp, patch.dict(
                        os.environ,
                        {
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                            "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_AGENT_TOKEN",
                            "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                            "ONECLI_AGENT_TOKEN": "test-onecli-agent-token",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    executor = ToolRuntimeExecutor(
                        policy_engine=PolicyEngine(allowed_http_hosts=["api.example.test"]),
                        run_store=JsonlToolRunStore(Path(tmp) / "tool_runs.jsonl"),
                    )
                    response = executor.run(
                        "agency.http.request",
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                        },
                        actor="user-runtime",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(response.verdict, "ok")
                    self.assertEqual(response.result["status_code"], 429)
                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    self.assertEqual(
                        [event.metadata.get("semanticType") for event in onecli_events],
                        ["onecli.http.request.started", "onecli.http.request.rate_limited"],
                    )
                    self.assertEqual(onecli_events[-1].metadata["status_code"], 429)
                    self.assertEqual(onecli_events[-1].metadata["verdict"], "warn")
                    self.assertEqual(onecli_events[-1].level.value, "warning")
                    self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(onecli_events[-1].metadata))
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_native_python_http_tool_denies_direct_auth_material_in_onecli_mode(self, mock_request):
        async def run_assertions():
            with patch.dict(
                    os.environ,
                    {
                        "ONECLI_ENABLED": "true",
                        "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                    },
                    clear=False,
            ):
                reset_settings_cache()
                tool = get_tool_catalog_specs()["agency.http.request"].tool_definition
                for header_name in sorted(ONECLI_BLOCKED_HEADER_NAMES):
                    with self.subTest(header_name=header_name):
                        with self.assertRaises(ValueError) as exc:
                            await ToolRegistry().execute(
                                tool,
                                {
                                    "url": "https://api.example.test/items",
                                    "method": "GET",
                                    "credential_mode": "onecli",
                                    "headers": {header_name: "raw-token"},
                                },
                                execution_id="execution-onecli-direct-header",
                            )
                        self.assertIn("credential-bearing headers", str(exc.exception))
                        self.assertIn(header_name, str(exc.exception))
                for query_name in sorted(ONECLI_BLOCKED_QUERY_PARAM_NAMES):
                    with self.subTest(query_name=query_name):
                        with self.assertRaises(ValueError) as exc:
                            await ToolRegistry().execute(
                                tool,
                                {
                                    "url": "https://api.example.test/items",
                                    "method": "GET",
                                    "credential_mode": "onecli",
                                    "query_params": {query_name: "raw-key"},
                                },
                                execution_id="execution-onecli-direct-query",
                            )
                        self.assertIn("credential-bearing query parameters", str(exc.exception))
                        self.assertIn(query_name, str(exc.exception))
            reset_settings_cache()

        asyncio.run(run_assertions())
        mock_request.assert_not_called()

    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_native_python_http_tool_includes_execution_context_in_onecli_metadata(self, mock_request):
        mock_request.return_value = _RequestsResponse()

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with patch.dict(
                        os.environ,
                        {
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                            "ONECLI_AGENT_TOKEN_SECRET_REF": "env://ONECLI_AGENT_TOKEN",
                            "ONECLI_ALLOW_GLOBAL_AGENT_TOKEN_FALLBACK": "true",
                            "ONECLI_AGENT_TOKEN": "test-onecli-agent-token",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    tool = get_tool_catalog_specs()["agency.http.request"].tool_definition
                    result = await ToolRegistry().execute(
                        tool,
                        {
                            "url": "https://api.example.test/items",
                            "method": "GET",
                            "credential_mode": "onecli",
                        },
                        execution_id="execution-onecli-http",
                        workflow_id="workflow-onecli-http",
                        task_id="task-onecli-http",
                        agent_id="agent-onecli-http",
                        tool_call_id="tool-call-onecli-http",
                    )
                    await asyncio.sleep(0)

                    self.assertEqual(result["status_code"], 200)
                    self.assertEqual(result["credential_mode"], "onecli")
                    self.assertEqual(result["onecli"]["target_host"], "api.example.test")
                    self.assertTrue(result["onecli"]["correlation_id"].startswith("onecli-http:"))
                    self.assertEqual(
                        result["onecli"]["agency_context"],
                        {
                            "execution_id": "execution-onecli-http",
                            "workflow_id": "workflow-onecli-http",
                            "task_id": "task-onecli-http",
                            "agent_id": "agent-onecli-http",
                            "tool_call_id": "tool-call-onecli-http",
                        },
                    )
                    self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(result))

                    kwargs = mock_request.call_args.kwargs
                    self.assertEqual(
                        kwargs["proxies"],
                        {
                            "http": "http://x:test-onecli-agent-token@onecli.local:10255",
                            "https": "http://x:test-onecli-agent-token@onecli.local:10255",
                        },
                    )
                    self.assertEqual(
                        kwargs["headers"]["X-Agency-OneCLI-Correlation-ID"],
                        result["onecli"]["correlation_id"],
                    )
                    self.assertEqual(kwargs["headers"]["X-Agency-Execution-ID"], "execution-onecli-http")
                    self.assertEqual(kwargs["headers"]["X-Agency-Workflow-ID"], "workflow-onecli-http")
                    self.assertEqual(kwargs["headers"]["X-Agency-Task-ID"], "task-onecli-http")
                    self.assertEqual(kwargs["headers"]["X-Agency-Agent-ID"], "agent-onecli-http")
                    self.assertEqual(kwargs["headers"]["X-Agency-Tool-Call-ID"], "tool-call-onecli-http")
                    self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(kwargs["headers"]))

                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    self.assertEqual(
                        [event.metadata.get("semanticType") for event in onecli_events],
                        ["onecli.http.request.started", "onecli.http.request.completed"],
                    )
                    for event in onecli_events:
                        self.assertEqual(event.metadata["correlation_id"], result["onecli"]["correlation_id"])
                        self.assertEqual(event.metadata["agency_context"]["execution_id"], "execution-onecli-http")
                        self.assertEqual(event.metadata["agency_context"]["workflow_id"], "workflow-onecli-http")
                        self.assertEqual(event.metadata["agency_context"]["task_id"], "task-onecli-http")
                        self.assertEqual(event.metadata["agency_context"]["agent_id"], "agent-onecli-http")
                        self.assertNotIn("env://ONECLI_AGENT_TOKEN", str(event.metadata))
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_native_python_http_tool_interpolates_connector_binding(self, mock_request):
        mock_request.return_value = _RequestsResponse()

        async def run_assertions():
            tool = get_tool_catalog_specs()["agency.http.request"].tool_definition
            connector_binding = {
                "provider": "discord-bot",
                "credential_id": "credential-discord-support",
                "purpose": "support_delivery",
                "target_scope": {
                    "guild_id": "guild-456",
                    "channel_id": "channel-123",
                },
                "identity_summary": "Support Discord / #triage",
            }
            result = await ToolRegistry().execute(
                tool,
                {
                    "url": "https://api.example.test/channels/{channel_id}/messages",
                    "method": "POST",
                    "headers": {"X-Connector-Provider": "{connector_provider}"},
                    "query_params": {
                        "credential_id": "{credential_id}",
                        "guild_id": "{target_scope[guild_id]}",
                    },
                    "body": {
                        "credential_id": "{connector_credential_id}",
                        "channel_id": "{channel_id}",
                        "purpose": "{connector_purpose}",
                    },
                },
                execution_id="execution-connector-binding-http",
                workflow_id="workflow-connector-binding-http",
                connector_binding=connector_binding,
            )

            self.assertEqual(result["status_code"], 200)
            kwargs = mock_request.call_args.kwargs
            self.assertEqual(kwargs["url"], "https://api.example.test/channels/channel-123/messages")
            self.assertEqual(kwargs["headers"]["X-Connector-Provider"], "discord-bot")
            self.assertEqual(
                kwargs["params"],
                {"credential_id": "credential-discord-support", "guild_id": "guild-456"},
            )
            self.assertEqual(
                kwargs["json"],
                {
                    "credential_id": "credential-discord-support",
                    "channel_id": "channel-123",
                    "purpose": "support_delivery",
                },
            )

        asyncio.run(run_assertions())

    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_http_interpolation_keeps_connector_binding_authoritative(self, mock_request):
        mock_request.return_value = _RequestsResponse()
        context = SimpleNamespace(
            connector_binding={
                "provider": "discord-bot",
                "credential_id": "credential-trusted",
                "target_scope": {"channel_id": "channel-trusted"},
            }
        )

        with self.assertRaisesRegex(ValueError, "cannot override connector context keys"):
            execute_custom_api(
                url="https://api.example.test/channels/{channel_id}/{credential_id}",
                method="GET",
                tool_context=context,
                channel_id="channel-attacker",
                credential_id="credential-attacker",
            )
        mock_request.assert_not_called()

    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_native_python_http_tool_denies_onecli_when_global_kill_switch_enabled(self, mock_request):
        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with patch.dict(
                        os.environ,
                        {
                            "ONECLI_ENABLED": "true",
                            "ONECLI_EXTERNAL_CALLS_DISABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    tool = get_tool_catalog_specs()["agency.http.request"].tool_definition
                    with self.assertRaises(ValueError) as exc:
                        await ToolRegistry().execute(
                            tool,
                            {
                                "url": "https://api.example.test/items",
                                "method": "GET",
                                "credential_mode": "onecli",
                            },
                            execution_id="execution-onecli-kill-switch",
                            workflow_id="workflow-onecli-kill-switch",
                        )
                    await asyncio.sleep(0)

                    self.assertIn("ONECLI_EXTERNAL_CALLS_DISABLED", str(exc.exception))
                    mock_request.assert_not_called()
                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    self.assertEqual(
                        [event.metadata.get("semanticType") for event in onecli_events],
                        ["onecli.http.request.denied"],
                    )
                    self.assertEqual(onecli_events[-1].metadata["verdict"], "deny")
                    self.assertIn("ONECLI_EXTERNAL_CALLS_DISABLED is true", onecli_events[-1].metadata["denial_reasons"])
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())

    @patch("app.core.outbound_http.socket.getaddrinfo")
    @patch("app.tools.implementations.http_integrations.requests.request")
    def test_native_python_http_tool_failed_event_is_deny_when_onecli_unavailable_in_production(
            self, mock_request, getaddrinfo):
        getaddrinfo.return_value = [(2, 1, 6, "", ("93.184.216.34", 443))]
        mock_request.side_effect = RuntimeError("onecli gateway unavailable")

        async def run_assertions():
            bus = RuntimeEventBus()
            set_default_runtime_event_bus(bus)
            subscriber = await bus.subscribe()
            try:
                with patch.dict(
                        os.environ,
                        {
                            "APP_ENV": "production",
                            "DATABASE_URL": "sqlite+aiosqlite:///:memory:",
                            "AGENCY_INTERNAL_API_KEY": "trusted-key",
                            "ONECLI_ENABLED": "true",
                            "ONECLI_GATEWAY_URL": "http://onecli.local:10255",
                            "TOOL_HTTP_ALLOWED_HOSTS": "api.example.test",
                        },
                        clear=False,
                ):
                    reset_settings_cache()
                    tool = get_tool_catalog_specs()["agency.http.request"].tool_definition
                    with self.assertRaises(RuntimeError) as exc:
                        await ToolRegistry().execute(
                            tool,
                            {
                                "url": "https://api.example.test/items",
                                "method": "GET",
                                "credential_mode": "onecli",
                            },
                            execution_id="execution-onecli-prod-fail-closed",
                            workflow_id="workflow-onecli-prod-fail-closed",
                        )
                    await asyncio.sleep(0)

                    self.assertIn("onecli gateway unavailable", str(exc.exception))
                    kwargs = mock_request.call_args.kwargs
                    self.assertEqual(
                        kwargs["proxies"],
                        {
                            "http": "http://onecli.local:10255",
                            "https": "http://onecli.local:10255",
                        },
                    )
                    events = []
                    while not subscriber.empty():
                        events.append(await subscriber.get())
                    onecli_events = [
                        event for event in events
                        if str(event.metadata.get("semanticType", "")).startswith("onecli.http.request.")
                    ]
                    self.assertEqual(
                        [event.metadata.get("semanticType") for event in onecli_events],
                        ["onecli.http.request.started", "onecli.http.request.failed"],
                    )
                    self.assertEqual(onecli_events[-1].metadata["verdict"], "deny")
                    self.assertTrue(onecli_events[-1].metadata["fail_closed"])
                    self.assertEqual(onecli_events[-1].level.value, "error")
            finally:
                reset_settings_cache()
                set_default_runtime_event_bus(None)

        asyncio.run(run_assertions())


if __name__ == "__main__":
    unittest.main()
