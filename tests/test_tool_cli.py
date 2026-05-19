from __future__ import annotations

import io
import json
import os
import subprocess
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.cli import main
from app.core.config import reset_settings_cache
from app.tools.runtime import JsonlToolRunStore


README_PATCH = """diff --git a/README.md b/README.md
--- a/README.md
+++ b/README.md
@@ -1,1 +1,1 @@
-hello
+hello cli
"""


def _create_workbook(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    workbook.save(path)
    return path


class ToolDiscoveryCliTests(unittest.TestCase):
    def tearDown(self) -> None:
        reset_settings_cache()

    def test_tool_list_json_returns_builtin_tools(self) -> None:
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = main(["tool", "list", "--json"])

        self.assertEqual(code, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["count"], 35)
        tool_ids = {item["id"] for item in payload["items"]}
        self.assertIn("agency.audio.transcribe", tool_ids)
        self.assertIn("agency.browser.open", tool_ids)
        self.assertIn("agency.command.run", tool_ids)

    def test_tool_describe_accepts_command_alias(self) -> None:
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = main(["tool", "describe", "browser open", "--json"])

        self.assertEqual(code, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["id"], "agency.browser.open")
        self.assertEqual(payload["command_alias"], "browser open")
        self.assertIn("input_schema", payload)
        self.assertIn("output_schema", payload)

    def test_tool_describe_accepts_pascal_case_name(self) -> None:
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = main(["tool", "describe", "Open Browser", "--json"])

        self.assertEqual(code, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["id"], "agency.browser.open")
        self.assertEqual(payload["name"], "open_browser")
        self.assertEqual(payload["display_name"], "Open Browser")

    def test_tool_describe_accepts_safe_call_name(self) -> None:
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = main(["tool", "describe", "OpenBrowser", "--json"])

        self.assertEqual(code, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["id"], "agency.browser.open")

    def test_tool_schema_can_emit_input_schema_only(self) -> None:
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = main(["tool", "schema", "agency.file.write-text", "--which", "input", "--json"])

        self.assertEqual(code, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["id"], "agency.file.write-text")
        self.assertIn("input_schema", payload)
        self.assertNotIn("output_schema", payload)
        self.assertEqual(payload["input_schema"]["properties"]["mode"]["enum"], ["write", "append"])

    def test_unknown_tool_returns_corrective_suggestions(self) -> None:
        stdout = io.StringIO()
        stderr = io.StringIO()
        with redirect_stdout(stdout), redirect_stderr(stderr):
            code = main(["tool", "describe", "browser opn"])

        self.assertEqual(code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertIn("Unknown tool: browser opn", stderr.getvalue())
        self.assertIn("agency.browser.open", stderr.getvalue())

    def test_tool_run_executes_contract_runtime_dry_run(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            repo = Path(tmp)
            subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True, text=True)
            (repo / "README.md").write_text("hello\n", encoding="utf-8")
            store_path = repo / "cli_tool_runs.jsonl"
            payload = {
                "repo": str(repo),
                "ref": "main",
                "changes": [{"path": "README.md", "patch": README_PATCH}],
                "dryRun": True,
            }

            stdout = io.StringIO()
            with (
                patch.dict(
                    os.environ,
                    {
                        "SANDBOX_EDIT_ALLOWED_REPOS": str(repo),
                        "TOOL_RUN_STORE_PATH": str(store_path),
                    },
                    clear=False,
                ),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(["tool", "run", "sandbox-edit", "--json", json.dumps(payload), "--output-json"])

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["filesChanged"][0]["path"], "README.md")
            self.assertTrue(response["signature"].startswith("sha256:"))
            self.assertEqual((repo / "README.md").read_text(encoding="utf-8"), "hello\n")
            records = JsonlToolRunStore(store_path).list_records()
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].actor, "cli")

    def test_tool_run_returns_policy_deny_response(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            payload = {
                "repo": "/tmp/not-allowed",
                "ref": "main",
                "changes": [{"path": ".env", "patch": "+OPENAI_API_KEY=sk-test"}],
                "dryRun": True,
            }
            stdout = io.StringIO()
            with (
                patch.dict(
                    os.environ,
                    {
                        "SANDBOX_EDIT_ALLOWED_REPOS": "",
                        "TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl"),
                    },
                    clear=False,
                ),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(["tool", "run", "sandbox-edit", "--json", json.dumps(payload), "--output-json"])

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "deny")
            self.assertTrue(response["errors"])

    def test_tool_run_rejects_invalid_json(self) -> None:
        stderr = io.StringIO()
        with redirect_stderr(stderr):
            code = main(["tool", "run", "sandbox-edit", "--json", "{not-json"])

        self.assertEqual(code, 1)
        self.assertIn("Invalid JSON payload", stderr.getvalue())

    def test_tool_run_rejects_non_contract_tool(self) -> None:
        stderr = io.StringIO()
        with redirect_stderr(stderr):
            code = main(["tool", "run", "agency.not-a-tool", "--json", "{}"])

        self.assertEqual(code, 1)
        self.assertIn("Tool is not contract-runnable", stderr.getvalue())
        self.assertIn("sandbox-edit", stderr.getvalue())

    @patch("app.tools.runtime.executor.open_browser")
    def test_tool_run_executes_contract_backed_browser_open(self, mock_open_browser) -> None:
        mock_open_browser.return_value = {
            "url": "https://example.test",
            "title": "Example",
            "runtime_root": "/tmp/browser-runtime",
            "message": "Browser started.",
        }
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.browser.open",
                        "--json",
                        json.dumps({"url": "https://example.test"}),
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status"], "ok")
            self.assertEqual(response["result"]["output"]["title"], "Example")
            self.assertTrue(response["signature"].startswith("sha256:"))
            mock_open_browser.assert_called_once()

    def test_tool_run_executes_contract_backed_tool_list(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(["tool", "run", "agency.tool.list", "--json", "{}", "--output-json"])

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["count"], 35)
            self.assertIn("agency.audio.transcribe", {item["id"] for item in response["result"]["items"]})
            self.assertIn("agency.tool.list", {item["id"] for item in response["result"]["items"]})

    def test_tool_run_executes_contract_backed_tool_get(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.tool.get",
                        "--json",
                        json.dumps({"tool_id": "agency.http.request"}),
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["tool"]["id"], "agency.http.request")

    def test_tool_run_marks_direct_proposal_tool_as_conversation_context_required(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.workflow.propose-create",
                        "--json",
                        json.dumps({"goal": "Create a workflow that drafts a report."}),
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "warn")
            self.assertEqual(response["result"]["status"], "requires_conversation_context")

    @patch("app.tools.runtime.executor.execute_custom_api")
    def test_tool_run_executes_contract_backed_http_request(self, mock_execute) -> None:
        mock_execute.return_value = {"status_code": 200, "response": {"ok": True}}
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(
                    os.environ,
                    {
                        "TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl"),
                        "TOOL_HTTP_ALLOWED_HOSTS": "api.example.test",
                    },
                    clear=False,
                ),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.http.request",
                        "--json",
                        json.dumps({"url": "https://api.example.test/items", "method": "GET"}),
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status_code"], 200)
            self.assertEqual(response["result"]["response"], {"ok": True})

    def test_tool_run_executes_contract_backed_command(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.command.run",
                        "--json",
                        json.dumps({"command": "printf 'cli-command\\n'", "mode": "bash", "timeout_seconds": 2}),
                        "--actor",
                        "approved/cli",
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status"], "ok")
            self.assertEqual(response["result"]["stdout"], "cli-command")

    def test_tool_run_denies_blocked_contract_backed_command(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.command.run",
                        "--json",
                        json.dumps({"command": "cat .env", "mode": "bash"}),
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "deny")
            self.assertIsNone(response["result"])
            self.assertIn("reading .env files is blocked", response["errors"][0])

    def test_tool_run_executes_contract_backed_file_write_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            stdout = io.StringIO()
            payload = {
                "base_folder": str(root),
                "filename": "cli.txt",
                "content": "cli file",
                "mode": "write",
            }
            with (
                patch.dict(
                    os.environ,
                    {
                        "TOOL_RUN_STORE_PATH": str(root / "tool_runs.jsonl"),
                        "TOOL_FILE_WRITE_ALLOWED_DIRS": str(root),
                    },
                    clear=False,
                ),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.file.write-text",
                        "--json",
                        json.dumps(payload),
                        "--actor",
                        "approved/cli",
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status"], "success")
            self.assertEqual((root / "cli.txt").read_text(encoding="utf-8"), "cli file")

    @patch("app.tools.implementations.documents.upload_to_s3")
    def test_tool_run_executes_contract_backed_markdown_to_word(self, mock_upload) -> None:
        mock_upload.return_value = {"uploaded_files": ["user_cli/workflow_reports/run_proc-cli/report.docx"]}
        with tempfile.TemporaryDirectory() as tmp:
            stdout = io.StringIO()
            payload = {
                "markdown_text": "# CLI Report\n\nBody",
                "filename": "report.docx",
                "img_directory": "reports",
                "process_id": "proc-cli",
                "run_by": "cli",
            }
            with (
                patch.dict(os.environ, {"TOOL_RUN_STORE_PATH": str(Path(tmp) / "tool_runs.jsonl")}, clear=False),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.document.markdown-to-word",
                        "--json",
                        json.dumps(payload),
                        "--actor",
                        "approved/cli",
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status"], "success")
            self.assertEqual(response["result"]["storage_uri"], "s3://mybucket/user_cli/workflow_reports/run_proc-cli/report.docx")

    def test_tool_run_executes_contract_backed_excel_text_writer(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = _create_workbook(root / "results.xlsx")
            text_path = root / "result.txt"
            text_path.write_text("cli spreadsheet", encoding="utf-8")
            stdout = io.StringIO()
            payload = {
                "sheet_name": "Sheet1",
                "excel_file_path": str(workbook_path),
                "text_file_path": str(text_path),
                "serial_number": 1,
                "header_title": "Notes",
            }
            with (
                patch.dict(
                    os.environ,
                    {
                        "TOOL_RUN_STORE_PATH": str(root / "tool_runs.jsonl"),
                        "TOOL_FILE_WRITE_ALLOWED_DIRS": str(root),
                    },
                    clear=False,
                ),
                redirect_stdout(stdout),
            ):
                reset_settings_cache()
                code = main(
                    [
                        "tool",
                        "run",
                        "agency.excel.write-text",
                        "--json",
                        json.dumps(payload),
                        "--actor",
                        "approved/cli",
                        "--output-json",
                    ]
                )

            self.assertEqual(code, 0)
            response = json.loads(stdout.getvalue())
            self.assertEqual(response["verdict"], "ok")
            self.assertEqual(response["result"]["status"], "success")
            self.assertEqual(load_workbook(workbook_path).active["A2"].value, "cli spreadsheet")


if __name__ == "__main__":
    unittest.main()
